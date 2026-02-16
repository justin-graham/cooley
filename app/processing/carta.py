"""Carta cap table parsing and tie-out comparison."""

import difflib
import logging
from typing import Any, Dict, List, Optional

from app import db
from app.processing.models import normalize_issue

logger = logging.getLogger(__name__)


def parse_carta_captable(xlsx_path: str) -> Dict[str, Any]:
    """Parse a Carta-exported cap table (.xlsx) into structured data."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    shareholders = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, headers = _find_header_row(ws)
        if header_row is None:
            continue

        name_col, shares_col, ownership_col = _detect_columns(headers)
        if name_col is None or shares_col is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            entry = _parse_shareholder_row(row, name_col, shares_col, ownership_col)
            if entry:
                shareholders.append(entry)

        if shareholders:
            break

    # Aggregate duplicates
    aggregated = {}
    for sh in shareholders:
        key = sh['name']
        if key not in aggregated:
            aggregated[key] = {'name': key, 'shares': 0.0, 'share_class': sh['share_class'], 'ownership_pct': 0.0}
        aggregated[key]['shares'] += sh['shares']

    shareholders = list(aggregated.values())
    total_shares = sum(sh['shares'] for sh in shareholders)
    if total_shares > 0:
        for sh in shareholders:
            sh['ownership_pct'] = round((sh['shares'] / total_shares) * 100, 2)

    return {'shareholders': shareholders, 'total_shares': total_shares}


def compare_cap_tables(carta_shareholders: List[Dict], generated_cap_table: List[Dict]) -> List[Dict]:
    """Compare Carta cap table against pipeline-generated cap table."""
    issues = []

    # Build normalized lookups with aggregation
    gen_by_name, gen_dups = _build_name_lookup(generated_cap_table, 'shareholder')
    carta_by_name, carta_dups = _build_name_lookup(carta_shareholders, 'name')

    for dup in sorted(gen_dups):
        issues.append({'severity': 'warning', 'category': 'Cap Table Tie-Out',
                       'description': f'Generated cap table has duplicate shareholder entries for "{gen_by_name[dup].get("shareholder", dup)}". Shares were aggregated before matching.'})
    for dup in sorted(carta_dups):
        issues.append({'severity': 'warning', 'category': 'Cap Table Tie-Out',
                       'description': f'Carta cap table has duplicate shareholder entries for "{carta_by_name[dup].get("name", dup)}". Shares were aggregated before matching.'})

    matched_gen = set()
    for carta_name, carta_entry in carta_by_name.items():
        match = _find_best_match(carta_name, gen_by_name.keys())
        if match is None:
            issues.append({'severity': 'warning', 'category': 'Cap Table Tie-Out',
                           'description': f'Shareholder "{carta_entry["name"]}" appears in Carta cap table ({carta_entry["shares"]:,.0f} shares) but was not found in source documents.'})
            continue

        matched_gen.add(match)
        gen_shares = float(gen_by_name[match].get('shares', 0))
        if abs(carta_entry['shares'] - gen_shares) > 0.5:
            issues.append({'severity': 'critical', 'category': 'Cap Table Tie-Out',
                           'description': f'Share mismatch for "{carta_entry["name"]}": Carta={carta_entry["shares"]:,.0f}, source={gen_shares:,.0f}.'})

    for gen_name, gen_entry in gen_by_name.items():
        if gen_name not in matched_gen:
            issues.append({'severity': 'warning', 'category': 'Cap Table Tie-Out',
                           'description': f'"{gen_entry["shareholder"]}" in source documents ({float(gen_entry.get("shares", 0)):,.0f} shares) not in Carta.'})

    # Total share comparison
    carta_total = sum(sh['shares'] for sh in carta_shareholders)
    gen_total = sum(float(e.get('shares', 0)) for e in generated_cap_table)
    if abs(carta_total - gen_total) > 0.5:
        issues.append({'severity': 'critical', 'category': 'Cap Table Tie-Out',
                       'description': f'Total share mismatch: Carta={carta_total:,.0f}, source={gen_total:,.0f}.'})

    return issues


def tieout_carta_captable(audit_id: str, captable_path: str):
    """Compare uploaded Carta cap table against generated cap table and store discrepancies."""
    carta_data = parse_carta_captable(captable_path)

    if not carta_data['shareholders']:
        logger.warning(f"Could not parse Carta cap table for audit {audit_id}")
        db.append_issues(audit_id, [normalize_issue({
            'severity': 'warning', 'category': 'Cap Table Tie-Out',
            'description': 'Uploaded cap table could not be parsed. Ensure it is a standard Carta export (.xlsx).',
        })])
        return

    audit = db.get_audit(audit_id)
    generated = audit.get('cap_table', []) if audit else []
    if not generated:
        return

    issues = compare_cap_tables(carta_data['shareholders'], generated)
    if issues:
        db.append_issues(audit_id, [normalize_issue(i) for i in issues])
        logger.info(f"Tie-out found {len(issues)} discrepancies for audit {audit_id}")
    else:
        db.append_issues(audit_id, [normalize_issue({
            'severity': 'info', 'category': 'Cap Table Tie-Out',
            'description': f'Carta matches source documents. All {len(carta_data["shareholders"])} shareholders verified.',
        })])


# --- Private helpers ---

def _find_header_row(ws):
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_vals = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[row_idx]]
        if any(kw in h for h in row_vals for kw in ['name', 'stakeholder']):
            return row_idx, row_vals
    return None, []


def _detect_columns(headers):
    name_col = shares_col = ownership_col = None
    for i, h in enumerate(headers):
        if name_col is None and any(kw in h for kw in ['stakeholder name', 'name']):
            name_col = i
        if shares_col is None and any(kw in h for kw in ['outstanding shares', 'quantity outstanding', 'shares outstanding']):
            shares_col = i
        if ownership_col is None and any(kw in h for kw in ['outstanding ownership', 'fully diluted ownership']):
            ownership_col = i
    if shares_col is None:
        for i, h in enumerate(headers):
            if 'common' in h and ('cs' in h or 'stock' in h):
                shares_col = i
                break
    return name_col, shares_col, ownership_col


def _parse_shareholder_row(row, name_col, shares_col, ownership_col):
    name = row[name_col] if name_col < len(row) else None
    shares_val = row[shares_col] if shares_col < len(row) else None
    if not name or not shares_val:
        return None

    name_str = str(name).strip()
    skip_keywords = ['total', 'totals', 'grand total', '']
    if name_str.lower() in skip_keywords:
        return None
    if any(kw in name_str.lower() for kw in ['total issued', 'total outstanding', 'options', 'shares available', 'fully diluted', 'percentage']):
        return None

    try:
        shares = float(str(shares_val).replace(',', '').replace('$', '')) if isinstance(shares_val, str) else float(shares_val)
    except (ValueError, TypeError):
        return None
    if shares <= 0:
        return None

    ownership = 0.0
    if ownership_col and ownership_col < len(row) and row[ownership_col]:
        try:
            ownership = float(str(row[ownership_col]).replace('%', ''))
        except (ValueError, TypeError):
            pass

    return {'name': name_str, 'shares': shares, 'share_class': 'Common Stock', 'ownership_pct': ownership}


def _normalize_name(name: str) -> str:
    return name.lower().strip().replace(',', '').replace('.', '') if name else ''


def _build_name_lookup(entries, name_key):
    by_name = {}
    duplicates = set()
    for entry in entries:
        name = _normalize_name(entry.get(name_key, ''))
        if not name:
            continue
        shares = float(entry.get('shares', 0) or 0)
        if name in by_name:
            duplicates.add(name)
            by_name[name]['shares'] = float(by_name[name].get('shares', 0) or 0) + shares
        else:
            by_name[name] = {**entry, 'shares': shares}
    return by_name, duplicates


def _find_best_match(name: str, candidates) -> Optional[str]:
    candidates = list(candidates)
    if not name or not candidates:
        return None
    if name in candidates:
        return name

    scored = [(c, difflib.SequenceMatcher(None, name, c).ratio()) for c in candidates]
    scored.sort(key=lambda x: x[1], reverse=True)
    best_name, best_score = scored[0]
    second_score = scored[1][1] if len(scored) > 1 else 0.0

    if best_score >= 0.92 and (best_score - second_score) >= 0.05:
        return best_name
    return None
