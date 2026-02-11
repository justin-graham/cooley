"""
AI Processing Pipeline - Orchestrates 3-pass document analysis using Claude.

Pass 1: Classify each document by type
Pass 2: Extract structured data from each document based on its type
Pass 3: Synthesize cross-document insights (timeline, cap table, issues)
"""

import os
import json
import logging
import time
from datetime import date, datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional, Tuple
from anthropic import Anthropic, APITimeoutError, APIError, RateLimitError
from pydantic import BaseModel, Field, ValidationError
from app import db, prompts

logger = logging.getLogger(__name__)


def format_text_with_paragraphs(text: str) -> str:
    """
    Format text with paragraph markers [¶1], [¶2], etc. for precise citation.

    Args:
        text: Raw document text

    Returns:
        Text formatted with paragraph markers for Claude to reference
    """
    if not text:
        return text

    # Split into paragraphs (skip empty lines)
    lines = text.split('\n')
    formatted_lines = []
    paragraph_num = 1
    current_para = []

    for line in lines:
        stripped = line.strip()

        if not stripped:
            # Empty line - save accumulated paragraph
            if current_para:
                para_text = ' '.join(current_para)
                if len(para_text) > 20:  # Substantive paragraphs only
                    formatted_lines.append(f"[¶{paragraph_num}] {para_text}")
                    paragraph_num += 1
                current_para = []
            continue

        current_para.append(stripped)

    # Don't forget last paragraph
    if current_para:
        para_text = ' '.join(current_para)
        if len(para_text) > 20:
            formatted_lines.append(f"[¶{paragraph_num}] {para_text}")

    return '\n\n'.join(formatted_lines)


def clean_text_for_db(text: str) -> str:
    """
    Clean text to remove characters that break PostgreSQL JSONB storage.

    Args:
        text: Input text string

    Returns:
        Cleaned text safe for PostgreSQL
    """
    if not isinstance(text, str):
        return text
    # Remove NULL bytes and other control characters that break PostgreSQL
    return text.replace('\x00', '').replace('\r', '\n')


def clean_document_dict(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Recursively clean all text fields in a document dictionary.

    Args:
        doc: Document dictionary

    Returns:
        Cleaned document dictionary
    """
    cleaned = {}
    for key, value in doc.items():
        if isinstance(value, str):
            cleaned[key] = clean_text_for_db(value)
        elif isinstance(value, dict):
            cleaned[key] = clean_document_dict(value)
        elif isinstance(value, list):
            cleaned[key] = [clean_document_dict(item) if isinstance(item, dict) else clean_text_for_db(item) if isinstance(item, str) else item for item in value]
        elif isinstance(value, Decimal):
            cleaned[key] = float(value)
        elif isinstance(value, (date, datetime)):
            cleaned[key] = value.isoformat()
        else:
            cleaned[key] = value
    return cleaned


# Initialize Claude client with 60 second timeout
client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"), timeout=60.0)


class ExtractedDataEnvelope(BaseModel):
    """Versioned schema for persisted documents.extracted_data payloads."""
    schema_version: str = "v1"
    category: str
    parse_status: str
    parse_error: Optional[str] = None
    extraction: Dict[str, Any] = Field(default_factory=dict)
    warnings: List[str] = Field(default_factory=list)


class IssueRecord(BaseModel):
    """Normalized issue object schema persisted in audits.issues."""
    severity: str
    category: str
    description: str
    source_doc: Optional[str] = None


def _normalize_compliance_status(value: Any, fallback: str = "WARNING") -> str:
    """
    Normalize event compliance statuses to DB-accepted enum values.
    """
    normalized = str(value or "").strip().upper()
    if normalized in {"VERIFIED", "WARNING", "CRITICAL"}:
        return normalized
    return fallback


def _severity_key(value: Any) -> str:
    sev = (value or "note").strip().lower()
    if sev in {"critical", "warning", "info", "note"}:
        return sev
    return "warning"


def normalize_issue(issue: Any) -> Dict[str, Any]:
    """
    Normalize issue shape for stable downstream rendering.
    """
    if isinstance(issue, str):
        normalized = {
            "severity": "note",
            "category": "General",
            "description": issue
        }
    elif isinstance(issue, dict):
        normalized = {
            "severity": _severity_key(issue.get("severity")),
            "category": str(issue.get("category") or "General"),
            "description": str(issue.get("description") or issue.get("message") or "Unspecified issue")
        }
        if issue.get("source_doc"):
            normalized["source_doc"] = str(issue.get("source_doc"))
    else:
        normalized = {
            "severity": "warning",
            "category": "System Error",
            "description": f"Unsupported issue payload: {type(issue).__name__}"
        }

    # Validate and return stable schema
    try:
        return IssueRecord(**normalized).model_dump(exclude_none=True)
    except ValidationError:
        return {
            "severity": "warning",
            "category": "System Error",
            "description": "Issue normalization failed"
        }


def _extract_doc_payload(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Build the extracted_data payload for one document with a stable schema.
    """
    metadata_keys = {
        "id",
        "document_id",
        "filename",
        "type",
        "text",
        "error",
        "category",
        "summary",
        "pdf_path",
        "text_spans",
        "parse_status",
        "parse_error",
        "preview_image",
        "preview_focus_y",
    }
    extraction = {k: v for k, v in doc.items() if k not in metadata_keys}
    warnings = []
    if doc.get("error"):
        warnings.append(str(doc.get("error")))
    if doc.get("parse_error") and doc.get("parse_error") not in warnings:
        warnings.append(str(doc.get("parse_error")))

    envelope = ExtractedDataEnvelope(
        category=doc.get("category", "Other"),
        parse_status=doc.get("parse_status", "success"),
        parse_error=doc.get("parse_error"),
        extraction=extraction,
        warnings=warnings
    )
    return envelope.model_dump(exclude_none=True)


def _build_enriched_documents(extractions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Build persisted document objects for audits.documents and documents table.
    """
    enriched_docs = []
    for doc in extractions:
        enriched = {
            "id": doc.get("id"),
            "filename": doc.get("filename"),
            "type": doc.get("type"),
            "category": doc.get("category", "Other"),
            "summary": doc.get("summary"),
            "text": doc.get("text", ""),
            "error": doc.get("error"),
            "parse_status": doc.get("parse_status", "error" if doc.get("error") else "success"),
            "parse_error": doc.get("parse_error") or doc.get("error"),
            "extracted_data": _extract_doc_payload(doc),
        }
        if doc.get("document_id"):
            enriched["document_id"] = doc.get("document_id")
        enriched_docs.append(enriched)
    return enriched_docs


def _scan_low_confidence(doc: Dict[str, Any]) -> List[str]:
    """
    Collect low-confidence warnings from nested extraction payloads.
    """
    warnings = []
    extracted = doc.get("extracted_data", {}).get("extraction", {})
    if not isinstance(extracted, dict):
        return warnings

    for value in extracted.values():
        if isinstance(value, dict):
            if value.get("low_confidence"):
                warnings.append(value.get("confidence_warning") or "Low confidence extraction")
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict) and item.get("low_confidence"):
                    warnings.append(item.get("confidence_warning") or "Low confidence extraction")
    return warnings


def _scan_extraction_errors(doc: Dict[str, Any]) -> List[str]:
    """
    Collect extraction failures from nested payloads.
    """
    failures = []
    extracted = doc.get("extracted_data", {}).get("extraction", {})
    if not isinstance(extracted, dict):
        return failures

    for key, value in extracted.items():
        if isinstance(value, dict) and value.get("error"):
            failures.append(f"{key}: {value.get('error')}")
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict) and item.get("error"):
                    failures.append(f"{key}: {item.get('error')}")
    return failures


def build_quality_report(
    documents: List[Dict[str, Any]],
    transactions: List[Dict[str, Any]],
    issues: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Build auditable quality report and decide if manual review is required.
    """
    report = {
        "schema_version": "v1",
        "document_count": len(documents),
        "parsed_successfully": 0,
        "parse_failures": 0,
        "extraction_failures": 0,
        "low_confidence_count": 0,
        "missing_approvals": 0,
        "critical_compliance_event_count": 0,
        "critical_issue_count": 0,
        "blocking_reasons": [],
        "warnings": []
    }

    for doc in documents:
        if doc.get("parse_status") in {"success", "partial"}:
            report["parsed_successfully"] += 1
        else:
            report["parse_failures"] += 1
            report["blocking_reasons"].append(
                f"Document parsing failed: {doc.get('filename', 'unknown')}"
            )

        low_confidence_warnings = _scan_low_confidence(doc)
        if low_confidence_warnings:
            report["low_confidence_count"] += len(low_confidence_warnings)
            report["warnings"].extend(low_confidence_warnings)
            report["blocking_reasons"].append(
                f"Low-confidence extraction requires review: {doc.get('filename', 'unknown')}"
            )

        extraction_failures = _scan_extraction_errors(doc)
        if extraction_failures:
            report["extraction_failures"] += len(extraction_failures)
            report["warnings"].extend(extraction_failures)
            report["blocking_reasons"].append(
                f"Extraction failed for one or more required fields: {doc.get('filename', 'unknown')}"
            )

    for tx in transactions:
        tx_type = (tx.get("event_type") or "").lower()
        requires_approval = tx_type in {"issuance", "repurchase", "option_grant"}
        missing_approval = requires_approval and not tx.get("approval_doc_id")
        if missing_approval:
            report["missing_approvals"] += 1
            report["blocking_reasons"].append(
                f"Missing approval evidence for {tx.get('event_type')} event on {tx.get('event_date')}"
            )

        summary = str(tx.get("summary") or "").strip().lower()
        if summary and any(token in summary for token in ["n/a", "none", "unknown", "null"]):
            report["blocking_reasons"].append(
                f"Unresolved summary placeholders found for event on {tx.get('event_date')}"
            )

        compliance_status = _normalize_compliance_status(tx.get("compliance_status"), fallback="WARNING")
        if compliance_status == "CRITICAL":
            report["critical_compliance_event_count"] += 1
            report["blocking_reasons"].append(
                f"Critical compliance gap for {tx.get('event_type')} event on {tx.get('event_date')}"
            )

    normalized_issues = [normalize_issue(issue) for issue in issues]
    report["critical_issue_count"] = sum(
        1 for issue in normalized_issues if issue.get("severity") == "critical"
    )
    if report["critical_issue_count"] > 0:
        report["blocking_reasons"].append(
            f"{report['critical_issue_count']} critical compliance issue(s) detected"
        )

    # Deduplicate while preserving order
    report["warnings"] = list(dict.fromkeys(report["warnings"]))
    report["blocking_reasons"] = list(dict.fromkeys(report["blocking_reasons"]))
    report["review_required"] = bool(report["blocking_reasons"])
    return report


# ============================================================================
# SHARE CLASS NORMALIZATION
# ============================================================================

# Map various share class names to canonical forms
SHARE_CLASS_ALIASES = {
    # Common stock variants
    'common': 'Common Stock',
    'common stock': 'Common Stock',
    'common shares': 'Common Stock',
    'class a common': 'Common Stock',
    'class a common stock': 'Common Stock',
    'ordinary shares': 'Common Stock',
    'ordinary stock': 'Common Stock',

    # Series Seed variants
    'series seed': 'Series Seed Preferred',
    'series seed preferred': 'Series Seed Preferred',
    'series seed preferred stock': 'Series Seed Preferred',
    'seed preferred': 'Series Seed Preferred',

    # Series A variants
    'series a': 'Series A Preferred',
    'series a preferred': 'Series A Preferred',
    'series a preferred stock': 'Series A Preferred',
    'series a-1': 'Series A Preferred',
    'series a-1 preferred': 'Series A Preferred',

    # Series B variants
    'series b': 'Series B Preferred',
    'series b preferred': 'Series B Preferred',
    'series b preferred stock': 'Series B Preferred',

    # Convertible instruments
    'safe': 'SAFE',
    'simple agreement for future equity': 'SAFE',
    'convertible note': 'Convertible Note',
    'convertible promissory note': 'Convertible Note',

    # Options
    'option': 'Option',
    'stock option': 'Option',
    'iso': 'Option',
    'nso': 'Option',
    'nqso': 'Option',
}


def normalize_share_class(share_class: str) -> str:
    """
    Normalize share class names to canonical forms for consistency.

    Args:
        share_class: Raw share class name from extraction

    Returns:
        Normalized canonical share class name
    """
    if not share_class:
        return 'Common Stock'  # Default to common

    # Normalize for lookup
    normalized = share_class.lower().strip()

    # Check alias map
    if normalized in SHARE_CLASS_ALIASES:
        return SHARE_CLASS_ALIASES[normalized]

    # If not in map, return original with proper capitalization
    return share_class.strip().title()


# ============================================================================
# HYBRID CLASSIFICATION: Keyword Pre-Scanning
# ============================================================================

# High-confidence keyword patterns for document classification
# Format: (pattern, category, summary_template)
KEYWORD_PATTERNS = [
    # CRITICAL: Order matters - most specific patterns first

    # Tax forms (very specific)
    (r'83\s*\(\s*b\s*\)', '83(b) Election', '83(b) election form'),

    # Equity instruments (specific patterns)
    (r'simple\s+agreement\s+for\s+future\s+equity|(?:^|\s)SAFE(?:\s|$)', 'SAFE', 'SAFE investment agreement'),

    # Stock certificates MUST be checked BEFORE general "certificate of" patterns
    (r'stock\s+certificate|certificate\s+(no\.?|number)\s*\d+', 'Stock Certificate', 'Stock certificate'),

    # Charter documents (order by specificity)
    (r'amended\s+and\s+restated\s+certificate\s+of\s+incorporation', 'Charter Document', 'Amended and Restated Certificate of Incorporation'),
    (r'certificate\s+of\s+(incorporation|formation)', 'Charter Document', 'Certificate of Incorporation'),
    (r'articles\s+of\s+incorporation', 'Charter Document', 'Articles of Incorporation'),
    (r'articles\s+of\s+organization', 'Charter Document', 'Articles of Organization (LLC)'),
    (r'articles\s+of\s+association', 'Charter Document', 'Articles of Association'),

    # Bylaws and operating agreements
    (r'amended\s+and\s+restated\s+bylaws', 'Charter Document', 'Amended and Restated Bylaws'),
    (r'bylaws', 'Charter Document', 'Corporate bylaws'),
    (r'operating\s+agreement', 'Charter Document', 'LLC Operating Agreement'),

    # Stock transactions (specific before general)
    (r'stock\s+purchase\s+agreement|restricted\s+stock\s+purchase', 'Stock Purchase Agreement', 'Stock purchase agreement'),
    (r'subscription\s+agreement', 'Stock Purchase Agreement', 'Subscription agreement'),
    (r'(share|stock)\s+repurchase\s+agreement', 'Share Repurchase Agreement', 'Share repurchase agreement'),

    # Governance documents
    (r'consent\s+of\s+(board|directors|stockholders)|written\s+consent', 'Board/Shareholder Minutes', 'Written consent document'),
    (r'minutes\s+of.*meeting|meeting\s+of\s+the\s+(board|directors)', 'Board/Shareholder Minutes', 'Board/shareholder meeting minutes'),

    # Shareholder agreements
    (r'investor\s+rights\s+agreement', 'Board/Shareholder Minutes', 'Investor rights agreement'),
    (r'(?:voting|stockholder|shareholder)s?\s+agreement', 'Board/Shareholder Minutes', 'Voting/shareholder agreement'),

    # Equity compensation
    (r'option\s+grant\s+(agreement|notice)|stock\s+option\s+agreement', 'Option Grant Agreement', 'Stock option grant agreement'),
    (r'equity\s+incentive\s+plan|\d+\s+stock\s+plan', 'Equity Incentive Plan', 'Equity incentive plan document'),

    # Financing documents
    (r'warrant\s+(agreement|certificate)', 'Stock Purchase Agreement', 'Warrant agreement'),
    (r'convertible\s+note|promissory\s+note', 'Convertible Note', 'Convertible promissory note'),
    (r'series\s+seed\s+preferred\s+stock', 'Stock Purchase Agreement', 'Series Seed financing agreement'),

    # Regulatory filings
    (r'form\s+d|sec\s+form\s+d|notice\s+of\s+exempt', 'Corporate Records', 'SEC Form D filing'),

    # Other corporate documents
    (r'indemnification\s+agreement', 'Indemnification Agreement', 'Director/officer indemnification agreement'),
    (r'proprietary\s+information.*agreement|PIIA', 'IP/Proprietary Info Agreement', 'Proprietary information and inventions agreement'),
    (r'employment\s+agreement|offer\s+letter', 'Employment Agreement', 'Employment agreement'),

    # Cap table
    (r'cap(?:italization)?\s+table|cap\s+table\s+summary|ownership\s+summary', 'Financial Statement', 'Capitalization table'),
]


def classify_by_keywords(text: str) -> tuple:
    """
    Attempt to classify document using keyword patterns.

    Args:
        text: Document text (first few thousand chars sufficient)

    Returns:
        Tuple of (category, summary) if confident match found, else (None, None)
    """
    import re

    # Use first 3000 chars for keyword matching (enough for titles/headers)
    sample = text[:3000].lower()

    for pattern, category, summary in KEYWORD_PATTERNS:
        if re.search(pattern, sample, re.IGNORECASE):
            logger.info(f"Keyword match found: '{pattern}' -> {category}")
            return (category, summary)

    return (None, None)


def call_claude(prompt: str, max_tokens: int = 2048, max_retries: int = 3) -> str:
    """
    Call Claude API with a prompt and return the response text.
    Retries with exponential backoff on rate limits and timeouts.

    Args:
        prompt: The prompt to send
        max_tokens: Maximum tokens in response
        max_retries: Maximum retry attempts for transient errors

    Returns:
        Response text from Claude

    Raises:
        APITimeoutError: If all retries exhausted on timeout
        RateLimitError: If all retries exhausted on rate limit
        APIError: If there's a non-retryable API error
    """
    for attempt in range(max_retries):
        try:
            message = client.messages.create(
                model="claude-sonnet-4-5-20250929",
                max_tokens=max_tokens,
                temperature=0,  # Deterministic outputs for consistency
                messages=[{
                    "role": "user",
                    "content": prompt
                }]
            )
            return message.content[0].text

        except (RateLimitError, APITimeoutError) as e:
            wait_time = 2 ** (attempt + 1)  # 2s, 4s, 8s
            if attempt < max_retries - 1:
                logger.warning(f"Claude API {type(e).__name__} (attempt {attempt + 1}/{max_retries}), retrying in {wait_time}s...")
                time.sleep(wait_time)
            else:
                logger.error(f"Claude API {type(e).__name__} after {max_retries} attempts: {e}")
                raise
        except APIError as e:
            logger.error(f"Claude API error: {e}")
            raise


def parse_json_response(response_text: str) -> Any:
    """
    Parse JSON from Claude's response with robust error handling.

    Handles markdown wrapping, explanatory text, truncation, and repairs common JSON errors.

    Args:
        response_text: Raw response from Claude

    Returns:
        Parsed JSON object (dict or list)
    """
    import re

    text = response_text.strip()

    # Remove markdown code block formatting if present
    if text.startswith("```json"):
        text = text[7:]
    if text.startswith("```"):
        text = text[3:]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    # Attempt 1: Try parsing the cleaned text directly
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Attempt 2: Extract JSON from surrounding text (non-greedy match)
    json_match = re.search(r'(\[[\s\S]*?\]|\{[\s\S]*?\})\s*$', text)
    if json_match:
        text = json_match.group(1)

    # Attempt 3: Standard parsing on extracted text
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        logger.warning(f"Initial JSON parse failed at line {e.lineno} col {e.colno}: {e.msg}. Attempting repair...")

        # Attempt 2: Fix trailing commas (common Claude mistake)
        try:
            repaired = re.sub(r',(\s*[\]}])', r'\1', text)
            result = json.loads(repaired)
            logger.info("JSON repaired successfully (removed trailing commas)")
            return result
        except json.JSONDecodeError:
            pass

        # Log failure details and re-raise
        # NOTE: We intentionally do NOT use raw_decode() to accept truncated JSON.
        # For a legal compliance tool, partial extraction data is worse than no data.
        # The calling extraction functions handle failures gracefully with try/except.
        logger.error(f"JSON repair failed. Error: {e}")
        logger.error(f"Problematic section: {text[max(0, e.pos-50):e.pos+50]}")
        raise


# ============================================================================
# PASS 1: CLASSIFICATION
# ============================================================================

def classify_document(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Classify a document into a type category.
    Uses keyword-based classification first, falls back to Claude if no match.

    Args:
        doc: Document dict with 'text' field

    Returns:
        Updated doc dict with 'category' and 'summary' fields
    """
    if doc.get('error'):
        # Skip documents that failed to parse
        doc['category'] = 'Other'
        doc['summary'] = 'Failed to parse document'
        return doc

    try:
        # Try keyword-based classification first (fast, free, accurate for obvious docs)
        category, summary = classify_by_keywords(doc['text'])

        if category:
            # High-confidence keyword match - skip Claude API call
            doc['category'] = category
            doc['summary'] = summary
            logger.info(f"Document classified by keywords: {doc.get('filename', 'unknown')} -> {category}")
            return doc

        # No keyword match - use Claude for nuanced classification
        text_sample = doc['text'][:10000]
        prompt = prompts.CLASSIFICATION_PROMPT.format(text=text_sample)
        response = call_claude(prompt, max_tokens=512)

        result = parse_json_response(response)
        doc['category'] = result.get('doc_type', 'Other')
        doc['summary'] = result.get('summary', 'No summary available')
        logger.info(f"Document classified by Claude: {doc.get('filename', 'unknown')} -> {doc['category']}")

    except Exception as e:
        # If classification fails, default to "Other"
        doc['category'] = 'Other'
        doc['summary'] = f'Classification failed: {str(e)}'

    return doc


# ============================================================================
# PASS 2: EXTRACTION (by document type)
# ============================================================================

def verify_extraction(source_text: str, extracted_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Verify that extracted values appear in the source text to catch hallucinations.
    Adds a 'verification' field with confidence score and warnings.

    Args:
        source_text: Original document text
        extracted_data: Data extracted by Claude

    Returns:
        Verification results with confidence score (0-100)
    """
    import re

    warnings = []
    verifications = 0
    total_checks = 0

    # Normalize text for comparison (lowercase, remove extra whitespace)
    normalized_text = ' '.join(source_text.lower().split())

    # Check numeric fields (shares, amounts, authorized_shares)
    for field in ['shares', 'amount', 'principal', 'authorized_shares', 'valuation_cap']:
        if field in extracted_data and extracted_data[field]:
            total_checks += 1
            value = extracted_data[field]

            # Try to find the number in text (with some flexibility for formatting)
            if isinstance(value, (int, float)):
                # Check various number formats: 10000, 10,000, 10000.00
                patterns = [
                    str(int(value)),  # Plain number
                    f"{int(value):,}",  # With commas
                    f"{value:.2f}",  # With decimals
                ]

                found = any(pattern.replace(',', '') in normalized_text.replace(',', '') for pattern in patterns)

                if found:
                    verifications += 1
                else:
                    warnings.append(f"{field}={value} not found in source text")

    # Check text fields (shareholder, investor, recipient, company_name)
    for field in ['shareholder', 'investor', 'recipient', 'company_name']:
        if field in extracted_data and extracted_data[field]:
            total_checks += 1
            value = str(extracted_data[field]).lower()

            # Check if the name/text appears in source (case-insensitive)
            if value in normalized_text:
                verifications += 1
            else:
                # Check if parts of the name appear (e.g., "John Smith" -> check "john" and "smith")
                name_parts = value.split()
                if len(name_parts) > 1 and all(part in normalized_text for part in name_parts):
                    verifications += 1
                else:
                    warnings.append(f"{field}='{extracted_data[field]}' not found in source text")

    # Check date fields (YYYY-MM-DD format)
    for field in ['date', 'incorporation_date', 'grant_date', 'meeting_date', 'maturity_date']:
        if field in extracted_data and extracted_data[field]:
            total_checks += 1
            date_str = extracted_data[field]

            # Try various date formats: 2023-01-15, January 15, 2023, 01/15/2023
            try:
                from datetime import datetime
                dt = datetime.strptime(date_str, '%Y-%m-%d')

                # Check if year, month, day appear in text
                year_found = str(dt.year) in normalized_text
                month_name = dt.strftime('%B').lower()  # e.g., "January"
                month_found = month_name in normalized_text or str(dt.month) in normalized_text
                day_found = str(dt.day) in normalized_text

                if year_found and (month_found or day_found):
                    verifications += 1
                else:
                    warnings.append(f"{field}={date_str} not clearly found in source text")
            except (ValueError, ImportError):
                # If date parsing fails, skip verification for this field
                total_checks -= 1

    # Calculate confidence score
    confidence = int((verifications / total_checks * 100)) if total_checks > 0 else 100

    return {
        'confidence_score': confidence,
        'verified_fields': verifications,
        'total_checks': total_checks,
        'warnings': warnings if warnings else None
    }


def extract_charter_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from Charter documents with verification."""
    try:
        prompt = prompts.CHARTER_EXTRACTION_PROMPT.format(text=doc['text'][:20000])
        response = call_claude(prompt, max_tokens=1024)
        result = parse_json_response(response)

        # Add source document reference
        result['source_doc'] = doc['filename']

        # Verify extraction against source text
        verification = verify_extraction(doc['text'][:20000], result)
        result['verification'] = verification

        if verification['confidence_score'] < 70:
            logger.warning(f"Low confidence charter extraction ({verification['confidence_score']}%) for {doc['filename']}: {verification.get('warnings')}")
            result['low_confidence'] = True
            result['confidence_warning'] = f"Low confidence extraction ({verification['confidence_score']}%) for {doc['filename']}. Manual review recommended."

        return result
    except Exception as e:
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_stock_data(doc: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extract equity issuances from Stock Purchase Agreements with verification."""
    try:
        # Format text with paragraph markers for citation
        formatted_text = format_text_with_paragraphs(doc['text'][:20000])
        prompt = prompts.STOCK_EXTRACTION_PROMPT.format(text=formatted_text)
        response = call_claude(prompt, max_tokens=2048)
        issuances = parse_json_response(response)

        # Add source document reference and verification to each issuance
        for issuance in issuances:
            issuance['source_doc'] = doc['filename']

            # Verify each issuance
            verification = verify_extraction(doc['text'][:20000], issuance)
            issuance['verification'] = verification

            if verification['confidence_score'] < 70:
                logger.warning(f"Low confidence stock extraction ({verification['confidence_score']}%) for {doc['filename']}: {verification.get('warnings')}")
                issuance['low_confidence'] = True
                issuance['confidence_warning'] = f"Low confidence extraction ({verification['confidence_score']}%) for {doc['filename']}. Manual review recommended."

        return issuances
    except Exception as e:
        return []


def extract_safe_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from SAFE documents with verification."""
    try:
        # Format text with paragraph markers for citation
        formatted_text = format_text_with_paragraphs(doc['text'][:15000])
        prompt = prompts.SAFE_EXTRACTION_PROMPT.format(text=formatted_text)
        response = call_claude(prompt, max_tokens=1024)
        result = parse_json_response(response)

        # Add source document reference
        result['source_doc'] = doc['filename']

        # Verify extraction
        verification = verify_extraction(doc['text'][:15000], result)
        result['verification'] = verification

        if verification['confidence_score'] < 70:
            logger.warning(f"Low confidence SAFE extraction ({verification['confidence_score']}%) for {doc['filename']}: {verification.get('warnings')}")
            result['low_confidence'] = True
            result['confidence_warning'] = f"Low confidence extraction ({verification['confidence_score']}%) for {doc['filename']}. Manual review recommended."

        return result
    except Exception as e:
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_convertible_note_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from Convertible Note documents with verification."""
    try:
        prompt = prompts.CONVERTIBLE_NOTE_EXTRACTION_PROMPT.format(text=doc['text'][:15000])
        response = call_claude(prompt, max_tokens=1024)
        result = parse_json_response(response)

        # Add source document reference
        result['source_doc'] = doc['filename']

        # Verify extraction
        verification = verify_extraction(doc['text'][:15000], result)
        result['verification'] = verification

        if verification['confidence_score'] < 70:
            logger.warning(f"Low confidence convertible note extraction ({verification['confidence_score']}%) for {doc['filename']}: {verification.get('warnings')}")
            result['low_confidence'] = True
            result['confidence_warning'] = f"Low confidence extraction ({verification['confidence_score']}%) for {doc['filename']}. Manual review recommended."

        return result
    except Exception as e:
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_board_minutes_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from Board/Shareholder Minutes."""
    try:
        prompt = prompts.BOARD_MINUTES_EXTRACTION_PROMPT.format(text=doc['text'][:15000])
        response = call_claude(prompt, max_tokens=1024)
        result = parse_json_response(response)
        # Add source document reference
        result['source_doc'] = doc['filename']
        return result
    except Exception as e:
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_option_grant_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from Option Grant Agreements."""
    try:
        # Format text with paragraph markers for citation
        formatted_text = format_text_with_paragraphs(doc['text'][:15000])
        prompt = prompts.OPTION_GRANT_EXTRACTION_PROMPT.format(text=formatted_text)
        response = call_claude(prompt, max_tokens=1024)
        result = parse_json_response(response)
        # Add source document reference
        result['source_doc'] = doc['filename']
        return result
    except Exception as e:
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_repurchase_data(doc: Dict[str, Any]) -> Dict[str, Any]:
    """Extract data from Share Repurchase Agreements."""
    try:
        # Format text with paragraph markers for citation
        formatted_text = format_text_with_paragraphs(doc['text'][:15000])
        prompt = prompts.SHARE_REPURCHASE_EXTRACTION_PROMPT.format(text=formatted_text)
        response = call_claude(prompt, max_tokens=1024)
        repurchase = parse_json_response(response)
        # Make shares negative to subtract from cap table
        if 'shares' in repurchase and isinstance(repurchase['shares'], (int, float)):
            repurchase['shares'] = -abs(repurchase['shares'])

        # Add source document reference
        repurchase['source_doc'] = doc['filename']

        # Log extracted repurchase for debugging
        logger.info(f"Extracted repurchase from {doc.get('filename', 'unknown')}: shareholder='{repurchase.get('shareholder')}', shares={repurchase.get('shares')}, class='{repurchase.get('share_class')}'")

        return repurchase
    except Exception as e:
        logger.error(f"Repurchase extraction failed for {doc.get('filename', 'unknown')}: {e}")
        return {'error': str(e), 'source_doc': doc['filename']}


def extract_by_type(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Route document to appropriate extraction function based on category.
    Also generates inline preview screenshots with color-coded highlights.

    Args:
        doc: Document with 'category', 'pdf_path', 'text_spans' fields

    Returns:
        Dictionary with extracted data, preview_image, and summary
    """
    category = doc.get('category', 'Other')

    # Route to appropriate extraction function
    if 'Charter' in category:
        result = {'charter_data': extract_charter_data(doc)}
    elif 'Stock Purchase' in category:
        result = {'stock_issuances': extract_stock_data(doc)}
    elif 'SAFE' in category:
        result = {'safe_data': extract_safe_data(doc)}
    elif 'Convertible Note' in category:
        result = {'convertible_note_data': extract_convertible_note_data(doc)}
    elif 'Minutes' in category:
        result = {'minutes_data': extract_board_minutes_data(doc)}
    elif 'Option Grant' in category:
        result = {'option_data': extract_option_grant_data(doc)}
    elif 'Repurchase' in category:
        result = {'repurchase_data': extract_repurchase_data(doc)}
    else:
        result = {}

    # Generate preview for equity documents with share/price data
    if any(keyword in category for keyword in ['Stock Purchase', 'Option Grant', 'SAFE', 'Repurchase']):
        pdf_path = doc.get('pdf_path')
        text_spans = doc.get('text_spans')
        doc_id = doc.get('id')

        # Pick the right extracted payload (handles list-based stock_issuances)
        extracted_data = None
        if 'Stock Purchase' in category and result.get('stock_issuances'):
            issuances = [i for i in result['stock_issuances'] if isinstance(i, dict)]
            if issuances:
                extracted_data = next(
                    (i for i in issuances if i.get('shares') or i.get('price_per_share') or i.get('shareholder')),
                    issuances[0]
                )
        else:
            data_key = next((k for k in result.keys() if k.endswith('_data')), None)
            extracted_data = result.get(data_key) if data_key else None

        if extracted_data and not extracted_data.get('error'):
            # Generate preview screenshot with colored highlights (if we have PDF + bboxes)
            if pdf_path and text_spans and doc_id:
                preview_base64, focus_y = generate_and_store_preview(
                    doc_id=doc_id,
                    pdf_path=pdf_path,
                    extracted_data=extracted_data,
                    text_spans=text_spans
                )

                if preview_base64:
                    result['preview_image'] = preview_base64
                if focus_y is not None:
                    result['preview_focus_y'] = focus_y
                else:
                    logger.info(f"No preview generated for {doc.get('filename')}: no highlight locations found")
            else:
                logger.info(
                    f"Skipping preview for {doc.get('filename')}: "
                    f"pdf_path={'yes' if pdf_path else 'no'}, text_spans={'yes' if text_spans else 'no'}, doc_id={'yes' if doc_id else 'no'}"
                )

            # Generate condensed summary (even if preview failed/missing)
            if 'Stock Purchase' in category:
                event_type = 'stock_issuance'
            elif 'Option Grant' in category:
                event_type = 'option_grant'
            elif 'SAFE' in category:
                event_type = 'safe'
            elif 'Repurchase' in category:
                event_type = 'repurchase'
            else:
                event_type = 'unknown'

            summary = generate_event_summary(extracted_data, event_type)
            result['summary'] = summary

            logger.info(f"Generated preview and summary for {doc.get('filename')}: {summary}")

    return result


# ============================================================================
# PHASE 2 HELPERS: SCREENSHOT GENERATION FOR INLINE PREVIEWS
# ============================================================================

def get_shareholder_color_backend(shareholder: str) -> str:
    """
    Backend version of getShareholderColor from frontend.
    Uses same hash logic to ensure color consistency with cap table.

    Args:
        shareholder: Shareholder name

    Returns:
        Hex color code (e.g., '#3B82F6')
    """
    COLORS = [
        '#3B82F6',  # Blue
        '#8B5CF6',  # Purple
        '#EC4899',  # Pink
        '#10B981',  # Green
        '#F59E0B',  # Amber
        '#EF4444',  # Red
        '#06B6D4',  # Cyan
        '#6366F1',  # Indigo
    ]

    # Hash shareholder name to color index
    hash_value = sum(ord(c) for c in shareholder)
    color_index = hash_value % len(COLORS)
    return COLORS[color_index]


def find_number_locations(
    extracted_data: Dict[str, Any],
    text_spans: List[Dict],
    doc_id: str
) -> List[Dict]:
    """
    Match extracted share counts and prices to their bounding boxes.

    Args:
        extracted_data: {shares: 1000000, price_per_share: 0.001, shareholder: "Jake Sortor"}
        text_spans: Bbox data from parse_pdf_with_bboxes()
        doc_id: Document UUID

    Returns:
        List of location dicts with page, bbox, color, data_type
    """
    locations = []

    # Extract target values
    shares = extracted_data.get('shares') or extracted_data.get('shares_issued')
    price = extracted_data.get('price_per_share')
    shareholder = extracted_data.get('shareholder') or extracted_data.get('recipient') or extracted_data.get('investor')

    if not shareholder:
        return locations

    # Get shareholder color (sync with cap table)
    color = get_shareholder_color_backend(shareholder)

    # Search for shares value in text_spans
    if isinstance(shares, (int, float)):
        shares = abs(shares)  # Use absolute value so repurchases still match document text

    if shares:
        # Format variations: "1000000", "1,000,000", "1000000.0"
        search_patterns = [
            str(int(shares)),
            f"{int(shares):,}",  # With commas
            f"{shares:.0f}",
        ]

        for pattern in search_patterns:
            for span in text_spans:
                if pattern in span['text'].replace(' ', ''):  # Remove spaces for matching
                    locations.append({
                        'doc_id': doc_id,
                        'page': span['page'],
                        'bbox': span['bbox'],
                        'text_value': span['text'],
                        'data_type': 'shares',
                        'shareholder': shareholder,
                        'color': color
                    })
                    break  # Found it, move on
            if any(loc['data_type'] == 'shares' for loc in locations):
                break  # Found a match, stop searching

    # Search for price value
    if price:
        search_patterns = [
            f"{price:.4f}",
            f"{price:.2f}",
            f"${price:.4f}",
            f"${price:.2f}",
        ]

        for pattern in search_patterns:
            for span in text_spans:
                if pattern in span['text']:
                    locations.append({
                        'doc_id': doc_id,
                        'page': span['page'],
                        'bbox': span['bbox'],
                        'text_value': span['text'],
                        'data_type': 'price',
                        'shareholder': shareholder,
                        'color': color
                    })
                    break
            if any(loc['data_type'] == 'price' for loc in locations):
                break

    return locations


def generate_preview_screenshot(
    pdf_path: str,
    locations: List[Dict],
    output_path: str,
    scale: float = 2.0
) -> tuple[str, float]:
    """
    Generate PNG screenshot of PDF page with highlighted numbers.

    Args:
        pdf_path: Path to PDF file
        locations: List of bbox locations to highlight
        output_path: Where to save PNG
        scale: Rendering scale (2.0 = high DPI)

    Returns:
        Path to generated PNG file
    """
    import fitz
    from PIL import Image, ImageDraw

    if not locations:
        logger.warning(f"No locations to highlight for {pdf_path}")
        return None, None

    doc = fitz.open(pdf_path)

    # Determine which page to render (first page with highlights)
    target_page = locations[0]['page']
    page = doc.load_page(target_page - 1)  # 0-indexed
    page_height = page.rect.height

    # Render page to pixmap
    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat)

    # Convert to PIL Image
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    draw = ImageDraw.Draw(img, 'RGBA')

    # Draw highlight rectangles
    for loc in locations:
        if loc['page'] == target_page:
            bbox = loc['bbox']
            color_hex = loc['color']

            # Scale coordinates
            x0, y0, x1, y1 = [coord * scale for coord in bbox]

            # Convert hex to RGB
            fill_color = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
            fill_color_alpha = fill_color + (50,)  # 50/255 = ~20% opacity

            # Draw semi-transparent filled rectangle
            draw.rectangle([x0, y0, x1, y1], fill=fill_color_alpha)

            # Draw border
            border_color = fill_color + (180,)  # 70% opacity border
            draw.rectangle([x0, y0, x1, y1], outline=border_color, width=2)

    # Save PNG
    img.save(output_path, 'PNG')
    doc.close()

    # Compute focus_y (center of union bbox normalized to page height)
    page_locations = [loc for loc in locations if loc['page'] == target_page]
    focus_y = None
    if page_locations:
        y0s = [loc['bbox'][1] for loc in page_locations]
        y1s = [loc['bbox'][3] for loc in page_locations]
        union_center = (min(y0s) + max(y1s)) / 2.0
        focus_y = union_center / page_height if page_height else None

    return output_path, focus_y


def generate_and_store_preview(
    doc_id: str,
    pdf_path: str,
    extracted_data: Dict,
    text_spans: List[Dict]
) -> str:
    """
    Generate preview screenshot and return base64 string.

    Args:
        doc_id: Document UUID
        pdf_path: Path to PDF file
        extracted_data: Extracted fields from document
        text_spans: Bbox data from parse_pdf_with_bboxes()

    Returns:
        (Base64-encoded PNG data (data:image/png;base64,...), focus_y) or (None, None) if no highlights
    """
    import base64
    import tempfile
    import os

    # Find number locations
    locations = find_number_locations(extracted_data, text_spans, doc_id)

    if not locations:
        logger.info(f"No numbers to highlight for doc {doc_id}, skipping preview generation")
        return None, None

    # Generate screenshot
    with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp:
        screenshot_path = tmp.name

    try:
        result_path, focus_y = generate_preview_screenshot(pdf_path, locations, screenshot_path)

        if not result_path:
            return None, None

        # Convert to base64 for database storage
        with open(result_path, 'rb') as img_file:
            img_data = img_file.read()
            base64_data = base64.b64encode(img_data).decode('utf-8')

        logger.info(f"Generated preview for doc {doc_id}: {len(locations)} highlights, {len(base64_data)} bytes")

        return f"data:image/png;base64,{base64_data}", focus_y

    except Exception as e:
        logger.error(f"Failed to generate preview for doc {doc_id}: {e}")
        return None, None

    finally:
        # Cleanup temp file
        if os.path.exists(screenshot_path):
            os.unlink(screenshot_path)


def generate_event_summary(extracted_data: Dict[str, Any], event_type: str) -> str:
    """
    Generate condensed event summary (20-30 words) for event cards.

    Format: "[Shareholder] received [shares] [share_class] at $[price] on [date]"

    Args:
        extracted_data: Extracted fields from document
        event_type: 'stock_issuance', 'option_grant', 'safe', 'repurchase', etc.

    Returns:
        Condensed summary string
    """
    shareholder = extracted_data.get('shareholder') or extracted_data.get('recipient') or extracted_data.get('investor') or 'Unknown party'
    shares = extracted_data.get('shares') or extracted_data.get('shares_issued')
    share_class = extracted_data.get('share_type') or extracted_data.get('share_class') or 'Common'
    price = extracted_data.get('price_per_share')
    date = extracted_data.get('date') or extracted_data.get('issuance_date') or extracted_data.get('grant_date') or 'Unknown date'

    # Use absolute value for display when dealing with negative deltas (e.g., repurchases)
    display_shares = abs(shares) if isinstance(shares, (int, float)) and event_type == 'repurchase' else shares

    # Format shares with commas
    shares_str = f"{int(display_shares):,}" if display_shares else "unspecified"

    # Format price
    price_str = f"${price:.4f}" if price else "unspecified price"

    # Build summary based on event type
    if event_type == 'stock_issuance':
        return f"{shareholder} received {shares_str} {share_class} shares at {price_str} per share on {date}"
    elif event_type == 'option_grant':
        return f"{shareholder} granted {shares_str} {share_class} options at {price_str} strike price on {date}"
    elif event_type == 'safe':
        amount = extracted_data.get('amount') or extracted_data.get('investment_amount')
        amount_str = f"${amount:,}" if amount else "unspecified amount"
        return f"{shareholder} invested {amount_str} via SAFE on {date}"
    elif event_type == 'repurchase':
        return f"Company repurchased {shares_str} {share_class} shares from {shareholder} on {date}"
    else:
        return f"{shareholder} - {shares_str} shares on {date}"


# ============================================================================
# PASS 3: SYNTHESIS
# ============================================================================

def build_timeline_programmatically(extractions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Build timeline using Python code (deterministic, guaranteed correctness).
    Extract events from structured data and sort chronologically.

    Args:
        extractions: List of documents with extracted data

    Returns:
        Sorted list of timeline events
    """
    events = []

    for doc in extractions:
        filename = doc.get('filename', 'unknown')

        # Charter/Incorporation events
        if 'charter_data' in doc:
            charter = doc['charter_data']
            if not charter.get('error') and charter.get('incorporation_date'):
                events.append({
                    'date': charter['incorporation_date'],
                    'event_type': 'formation',
                    'description': f"Company incorporated: {charter.get('company_name', 'Unknown Company')}",
                    'source_docs': [filename]
                })

        # Stock issuance events
        if 'stock_issuances' in doc:
            for issuance in doc['stock_issuances']:
                if issuance.get('date') and issuance.get('shareholder') and issuance.get('shares'):
                    events.append({
                        'date': issuance['date'],
                        'event_type': 'stock_issuance',
                        'description': f"{issuance['shareholder']} received {issuance['shares']:,} shares of {issuance.get('share_class', 'stock')}",
                        'source_docs': [filename]
                    })

        # SAFE events
        if 'safe_data' in doc:
            safe = doc['safe_data']
            if not safe.get('error') and safe.get('date') and safe.get('investor'):
                events.append({
                    'date': safe['date'],
                    'event_type': 'financing',
                    'description': f"SAFE investment by {safe['investor']} for ${safe.get('amount', 0):,}",
                    'source_docs': [filename]
                })

        # Convertible Note events
        if 'convertible_note_data' in doc:
            note = doc['convertible_note_data']
            if not note.get('error') and note.get('date') and note.get('investor'):
                interest = note.get('interest_rate', 0)
                maturity = note.get('maturity_date', 'unspecified maturity')
                events.append({
                    'date': note['date'],
                    'event_type': 'financing',
                    'description': f"Convertible note from {note['investor']} for ${note.get('principal', 0):,} at {interest}% interest (maturity: {maturity})",
                    'source_docs': [filename]
                })

        # Board meeting events
        if 'minutes_data' in doc:
            minutes = doc['minutes_data']
            if not minutes.get('error') and minutes.get('meeting_date'):
                decisions = minutes.get('key_decisions', [])
                decisions_str = '; '.join(decisions[:2]) if decisions else 'corporate actions discussed'
                events.append({
                    'date': minutes['meeting_date'],
                    'event_type': 'board_action',
                    'description': f"{minutes.get('meeting_type', 'Meeting')}: {decisions_str}",
                    'source_docs': [filename]
                })

        # Option grant events
        if 'option_data' in doc:
            option = doc['option_data']
            if not option.get('error') and option.get('grant_date') and option.get('recipient'):
                events.append({
                    'date': option['grant_date'],
                    'event_type': 'option_grant',
                    'description': f"Option grant to {option['recipient']} for {option.get('shares', 0):,} shares",
                    'source_docs': [filename]
                })

        # Repurchase events
        if 'repurchase_data' in doc:
            repurchase = doc['repurchase_data']
            if not repurchase.get('error') and repurchase.get('date') and repurchase.get('shareholder'):
                # Handle None shares values
                shares = repurchase.get('shares') or 0
                shares = abs(shares) if isinstance(shares, (int, float)) else 0

                events.append({
                    'date': repurchase['date'],
                    'event_type': 'repurchase',
                    'description': f"Company repurchased {shares:,} shares from {repurchase['shareholder']}" if shares > 0 else f"Share repurchase transaction with {repurchase['shareholder']}",
                    'source_docs': [filename]
                })

    # Sort by date (chronological order)
    events.sort(key=lambda x: x.get('date', ''))

    logger.info(f"Built timeline programmatically: {len(events)} events")
    return events


def synthesize_timeline(extractions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Generate a chronological timeline from all extracted data.
    Uses programmatic (code-based) timeline as primary, optionally enhances with AI.

    Args:
        extractions: List of documents with extracted data

    Returns:
        Sorted list of timeline events
    """
    # Build deterministic timeline first (guaranteed correctness)
    programmatic_timeline = build_timeline_programmatically(extractions)

    # If we have a good timeline, return it (no need for AI enhancement)
    if len(programmatic_timeline) >= 3:
        logger.info(f"Using programmatic timeline ({len(programmatic_timeline)} events)")
        return programmatic_timeline

    # For sparse timelines, optionally enhance with Claude (can add context/narrative)
    try:
        extractions_json = json.dumps(extractions, indent=2)
        prompt = prompts.TIMELINE_SYNTHESIS_PROMPT.format(extractions_json=extractions_json)

        # Call Claude - if rate limited, return programmatic timeline
        response = call_claude(prompt, max_tokens=4096)
        ai_timeline = parse_json_response(response)

        # Sort by date
        ai_timeline.sort(key=lambda x: x.get('date', ''))

        logger.info(f"Using AI-enhanced timeline ({len(ai_timeline)} events)")
        return ai_timeline

    except RateLimitError as e:
        logger.warning(f"Timeline synthesis rate limited, using programmatic timeline: {e}")
        print(f"WARNING: Timeline AI enhancement rate limited - using programmatic timeline ({len(programmatic_timeline)} events)")
        return programmatic_timeline
    except Exception as e:
        logger.error(f"Timeline synthesis failed, using programmatic timeline: {e}", exc_info=True)
        print(f"ERROR: Timeline AI enhancement failed - using programmatic timeline ({len(programmatic_timeline)} events)")
        return programmatic_timeline


def build_raw_cap_table(equity_data: List[Dict[str, Any]], issues: List[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """
    Build a basic cap table from raw equity data without AI synthesis.
    Uses share class normalization and Decimal arithmetic for precision.

    Args:
        equity_data: List of equity issuances
        issues: Optional list to append data integrity issues to

    Returns:
        List of cap table entries with calculated ownership percentages
    """
    from decimal import ROUND_HALF_UP

    # Aggregate by shareholder and normalized share class using Decimal for precision
    aggregated = {}
    for item in equity_data:
        shareholder = item.get('shareholder') or item.get('investor') or item.get('recipient') or 'Unknown'
        shares = item.get('shares', 0)
        raw_share_class = item.get('share_class') or item.get('type', 'Common Stock')

        # Normalize share class for consistency
        share_class = normalize_share_class(raw_share_class)

        key = (shareholder, share_class)
        if key not in aggregated:
            aggregated[key] = Decimal('0')

        # Add shares using Decimal for precision
        if isinstance(shares, (int, float, Decimal)):
            aggregated[key] += Decimal(str(shares))

    # Separate positions by sign and flag data integrity issues
    negative_positions = {k: v for k, v in aggregated.items() if v < 0}
    zero_positions = {k: v for k, v in aggregated.items() if v == 0}
    positive_positions = {k: v for k, v in aggregated.items() if v > 0}

    # Log and flag fully-repurchased shareholders (net zero)
    for (shareholder, share_class), _ in zero_positions.items():
        logger.info(f"Full repurchase: {shareholder} ({share_class}) has 0 net shares")
        if issues is not None:
            issues.append({
                'severity': 'info',
                'category': 'Cap Table',
                'description': f"{shareholder} has 0 net shares ({share_class}) after repurchase. Shareholder fully exited."
            })

    # Flag negative positions as data integrity errors
    for (shareholder, share_class), shares in negative_positions.items():
        logger.error(f"Data integrity: {shareholder} ({share_class}) has {shares} net shares (negative)")
        if issues is not None:
            issues.append({
                'severity': 'critical',
                'category': 'Data Integrity',
                'description': f"{shareholder} has {shares} net shares ({share_class}). More shares repurchased than issued — possible missing issuance document."
            })

    # Only include positive positions in cap table
    aggregated = positive_positions

    # Calculate total shares for ownership percentage using Decimal
    total_shares = sum(aggregated.values())

    # Convert to cap table format with ownership %
    TWO_PLACES = Decimal('0.01')
    cap_table = [
        {
            'shareholder': shareholder,
            'shares': float(shares),
            'share_class': share_class,
            'ownership_pct': float((shares / total_shares * 100).quantize(TWO_PLACES, rounding=ROUND_HALF_UP)) if total_shares > 0 else 0.0
        }
        for (shareholder, share_class), shares in aggregated.items()
    ]

    # Sort by ownership percentage descending
    cap_table.sort(key=lambda x: x['ownership_pct'], reverse=True)

    # Adjust largest shareholder to ensure total = exactly 100.00%
    if cap_table:
        calculated_total = sum(entry['ownership_pct'] for entry in cap_table)
        if calculated_total != 100.0:
            adjustment = round(100.0 - calculated_total, 2)
            cap_table[0]['ownership_pct'] = round(cap_table[0]['ownership_pct'] + adjustment, 2)
            logger.info(f"Adjusted largest cap table entry by {adjustment}% to ensure 100% total")

    return cap_table


def synthesize_cap_table(extractions: List[Dict[str, Any]]):
    """
    Generate a cap table from equity issuance data.
    Uses programmatic (code-based) aggregation as primary method.

    Args:
        extractions: List of documents with extracted data

    Returns:
        Tuple of (cap_table, issues) where cap_table is a list of entries
        and issues is a list of data integrity issues found during calculation
    """
    # Collect all equity-related extractions
    equity_data = []

    for doc in extractions:
        if 'stock_issuances' in doc:
            equity_data.extend(doc['stock_issuances'])
        if 'safe_data' in doc:
            safe = doc['safe_data']
            if not safe.get('error'):
                equity_data.append({
                    'shareholder': safe.get('investor'),
                    'amount': safe.get('amount'),
                    'type': 'SAFE',
                    'date': safe.get('date')
                })
        if 'convertible_note_data' in doc:
            note = doc['convertible_note_data']
            if not note.get('error'):
                equity_data.append({
                    'shareholder': note.get('investor'),
                    'amount': note.get('principal'),
                    'type': 'Convertible Note',
                    'date': note.get('date')
                })
        # Note: Options are excluded from cap table as they represent rights to purchase,
        # not actual ownership. They should only be shown in fully-diluted calculations.
        if 'repurchase_data' in doc:
            repurchase = doc['repurchase_data']
            if not repurchase.get('error'):
                shares = repurchase.get('shares')
                shareholder = repurchase.get('shareholder')
                share_class = repurchase.get('share_class', 'Common Stock')

                # If shares not extracted from repurchase doc, infer from original issuance
                if shares is None and shareholder:
                    logger.info(f"Repurchase shares=None for {shareholder}, attempting to infer from issuances...")

                    # Find matching shareholder's stock issuances
                    matching_issuances = [
                        item for item in equity_data
                        if (item.get('shareholder') == shareholder and
                            item.get('shares') and
                            isinstance(item.get('shares'), (int, float, Decimal)) and
                            item.get('shares') > 0)
                    ]

                    if matching_issuances:
                        # Sum all matching issuances for this shareholder
                        total_shares = sum(item['shares'] for item in matching_issuances)
                        shares = total_shares
                        logger.info(f"Inferred repurchase amount for {shareholder}: {shares} shares (from {len(matching_issuances)} issuances)")
                    else:
                        logger.warning(f"Could not infer repurchase shares for {shareholder} - no matching issuances found")

                # Add repurchase with negative shares to subtract from cap table
                if shares and isinstance(shares, (int, float, Decimal)):
                    equity_data.append({
                        'shareholder': shareholder,
                        'shares': -float(abs(shares)),  # Make negative to subtract (convert to float)
                        'share_class': share_class,
                        'date': repurchase.get('date')
                    })
                    logger.info(f"Added repurchase: {shareholder} -{float(abs(shares))} shares")

    if not equity_data:
        return [], []

    # Collect data integrity issues from cap table calculation
    cap_table_issues = []

    # Build cap table programmatically (deterministic, no math errors)
    programmatic_cap_table = build_raw_cap_table(equity_data, issues=cap_table_issues)
    logger.info(f"Built cap table programmatically: {len(programmatic_cap_table)} entries, {len(cap_table_issues)} issues")

    # Return programmatic cap table and any data integrity issues found
    return programmatic_cap_table, cap_table_issues


def check_deterministic_issues(documents: List[Dict[str, Any]], cap_table: List[Dict[str, Any]], timeline: List[Dict[str, Any]], extractions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Check for critical compliance issues using deterministic (code-based) rules.
    These checks are 100% reliable and don't depend on AI interpretation.

    Args:
        documents: List of classified documents
        cap_table: Generated cap table
        timeline: Generated timeline
        extractions: Raw extraction data with charter info

    Returns:
        List of issues found by deterministic checks
    """
    issues = []

    # Get document categories for analysis
    doc_categories = [d.get('category', '') for d in documents]

    # CRITICAL: Missing Charter Document
    has_charter = any('Charter' in cat for cat in doc_categories)
    if not has_charter:
        issues.append({
            'severity': 'critical',
            'category': 'Missing Document',
            'description': 'No Certificate of Incorporation or Charter Document found. This is required to establish the company\'s legal existence and authorized shares.'
        })

    # CRITICAL: Stock issuances without board approval
    has_stock_issuances = any('Stock Purchase' in cat or 'Stock Certificate' in cat for cat in doc_categories)
    has_board_consent = any('Minutes' in cat or 'Board' in cat for cat in doc_categories)

    if has_stock_issuances and not has_board_consent:
        issues.append({
            'severity': 'critical',
            'category': 'Equity Compliance',
            'description': 'Stock issuances found but no Board Minutes or Written Consents documenting approval. All stock issuances must be authorized by the board.'
        })

    # CRITICAL: Issued shares exceed authorized shares
    # Extract authorized shares from charter
    authorized_shares = None
    for ext in extractions:
        if 'charter_data' in ext:
            charter = ext['charter_data']
            if not charter.get('error') and charter.get('authorized_shares'):
                try:
                    authorized_shares = int(charter['authorized_shares'])
                    break
                except (ValueError, TypeError):
                    pass

    if authorized_shares and cap_table:
        # Calculate total issued shares from cap table
        total_issued = sum(
            float(entry.get('shares', 0))
            for entry in cap_table
            if isinstance(entry.get('shares'), (int, float, Decimal))
        )

        if total_issued > authorized_shares:
            issues.append({
                'severity': 'critical',
                'category': 'Cap Table Integrity',
                'description': f'Issued shares ({total_issued:,}) exceed authorized shares ({authorized_shares:,}). Company must amend charter to increase authorized shares before these issuances are valid.'
            })

    # WARNING: Founder stock without 83(b) elections
    has_founder_stock = any('Stock Purchase' in cat for cat in doc_categories)
    has_83b = any('83(b)' in cat for cat in doc_categories)

    if has_founder_stock and not has_83b:
        issues.append({
            'severity': 'warning',
            'category': 'Equity Compliance',
            'description': 'Stock Purchase Agreements found but no 83(b) election forms. Founders who received restricted stock should file 83(b) elections within 30 days to avoid adverse tax consequences.'
        })

    # WARNING: Few board meetings over long period
    if timeline:
        # Extract date range from timeline
        dates = [event.get('date', '') for event in timeline if event.get('date')]
        if dates:
            dates.sort()
            first_date = dates[0]
            last_date = dates[-1]

            # Calculate years between first and last event
            try:
                from datetime import datetime
                first = datetime.strptime(first_date, '%Y-%m-%d')
                last = datetime.strptime(last_date, '%Y-%m-%d')
                years = (last - first).days / 365.25

                # Count board meetings/consents
                board_meetings = sum(1 for d in documents if 'Minutes' in d.get('category', '') or 'Board' in d.get('category', ''))

                if years >= 3 and board_meetings < 3:
                    issues.append({
                        'severity': 'warning',
                        'category': 'Board Governance',
                        'description': f'Company has {years:.1f} years of history but only {board_meetings} documented board meeting(s)/consent(s). Regular board meetings are important for proper governance.'
                    })
            except (ValueError, ImportError):
                pass  # Skip if date parsing fails

    # NOTE: Missing option plan for option grants
    has_option_grants = any('Option Grant' in cat for cat in doc_categories)
    has_option_plan = any('Equity Incentive Plan' in cat or 'Stock Plan' in cat for cat in doc_categories)

    if has_option_grants and not has_option_plan:
        issues.append({
            'severity': 'note',
            'category': 'Equity Compliance',
            'description': 'Option Grant Agreements found but no Equity Incentive Plan document. Options should be granted under a board-approved plan.'
        })

    # CRITICAL: Missing share counts in repurchase agreements
    for ext in extractions:
        if 'repurchase_data' in ext:
            repurchase = ext['repurchase_data']
            filename = ext.get('filename', 'unknown')

            if not repurchase.get('error'):
                shareholder = repurchase.get('shareholder')
                date = repurchase.get('date')
                shares = repurchase.get('shares')

                # Flag if share count is missing
                if shares is None and shareholder and date:
                    issues.append({
                        'severity': 'critical',
                        'category': 'Missing Data',
                        'description': f"Repurchase from {shareholder} on {date} missing share count. Document: {filename}. Manual review required to determine repurchased share amount.",
                        'source_doc': filename
                    })

    # ========================================================================
    # PHASE C: Enhanced Issue Detection Rules
    # ========================================================================

    # CHRONOLOGICAL INTEGRITY: Formation must be earliest event
    if timeline:
        from datetime import datetime

        formation_date = None
        earliest_non_formation_date = None
        earliest_non_formation_event = None

        for event in timeline:
            event_date = event.get('date', '')
            event_type = event.get('event_type', '')

            if not event_date:
                continue

            try:
                dt = datetime.strptime(event_date, '%Y-%m-%d')

                if event_type == 'formation':
                    if formation_date is None or dt < formation_date:
                        formation_date = dt
                else:
                    if earliest_non_formation_date is None or dt < earliest_non_formation_date:
                        earliest_non_formation_date = dt
                        earliest_non_formation_event = event
            except ValueError:
                continue

        # Check if any events predate formation
        if formation_date and earliest_non_formation_date:
            if earliest_non_formation_date < formation_date:
                issues.append({
                    'severity': 'critical',
                    'category': 'Chronological Integrity',
                    'description': f"Event dated {earliest_non_formation_date.strftime('%Y-%m-%d')} ({earliest_non_formation_event.get('event_type', 'unknown')}) predates company formation ({formation_date.strftime('%Y-%m-%d')}). All corporate actions must occur after incorporation."
                })

    # OPTION POOL VALIDATION: Compare grants to declared pool size
    # Look for equity incentive plan to find pool size
    option_pool_size = None
    for ext in extractions:
        if 'Equity Incentive Plan' in ext.get('category', '') or 'Stock Plan' in ext.get('category', ''):
            # Try to extract pool size from document
            text = ext.get('text', '').lower()
            import re
            # Look for patterns like "reserve 1,000,000 shares" or "option pool of 2000000"
            pool_patterns = [
                r'(?:reserve|pool|allocated?|set aside)\s+(?:of\s+)?(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*(?:shares|options)',
                r'(\d{1,3}(?:,\d{3})*(?:\.\d+)?)\s*shares?\s+(?:reserved|allocated|in\s+the\s+pool)',
            ]
            for pattern in pool_patterns:
                match = re.search(pattern, text)
                if match:
                    try:
                        option_pool_size = int(match.group(1).replace(',', ''))
                        break
                    except (ValueError, AttributeError):
                        pass
            if option_pool_size:
                break

    # Calculate total option grants
    total_option_grants = 0
    for ext in extractions:
        if 'option_data' in ext:
            option = ext['option_data']
            if not option.get('error') and option.get('shares'):
                try:
                    total_option_grants += int(option['shares'])
                except (ValueError, TypeError):
                    pass

    # Check if grants exceed pool
    if option_pool_size and total_option_grants > option_pool_size:
        issues.append({
            'severity': 'critical',
            'category': 'Option Pool Integrity',
            'description': f"Total option grants ({total_option_grants:,} shares) exceed declared option pool size ({option_pool_size:,} shares). Pool must be increased before additional grants are valid."
        })
    elif option_pool_size and total_option_grants > 0:
        utilization = (total_option_grants / option_pool_size) * 100
        if utilization > 90:
            issues.append({
                'severity': 'warning',
                'category': 'Option Pool Integrity',
                'description': f"Option pool is {utilization:.1f}% utilized ({total_option_grants:,} of {option_pool_size:,} shares). Consider increasing pool size before next grant."
            })

    # MISSING DOCUMENT INFERENCE: Check board minutes for referenced documents
    referenced_docs = []
    for ext in extractions:
        if 'minutes_data' in ext:
            minutes = ext['minutes_data']
            if not minutes.get('error'):
                decisions = minutes.get('key_decisions', [])
                for decision in decisions:
                    decision_lower = str(decision).lower()
                    # Look for references to documents that should exist
                    if 'option grant' in decision_lower and 'approved' in decision_lower:
                        referenced_docs.append(('Option Grant Agreement', decision))
                    if 'stock issuance' in decision_lower or 'issue shares' in decision_lower:
                        referenced_docs.append(('Stock Purchase Agreement', decision))
                    if 'safe' in decision_lower and 'approved' in decision_lower:
                        referenced_docs.append(('SAFE', decision))

    # Check if referenced documents exist
    for doc_type, decision in referenced_docs:
        has_doc = any(doc_type in d.get('category', '') for d in documents)
        if not has_doc:
            issues.append({
                'severity': 'warning',
                'category': 'Missing Document',
                'description': f"Board minutes reference '{decision[:80]}...' but no {doc_type} document found. Document may be missing from the upload."
            })

    # LOW CONFIDENCE EXTRACTIONS: Flag for manual review
    for ext in extractions:
        # Inspect all nested extracted payloads
        for key, value in ext.items():
            if isinstance(value, dict) and value.get('low_confidence'):
                issues.append({
                    'severity': 'warning',
                    'category': 'Extraction Quality',
                    'description': value.get(
                        'confidence_warning',
                        f"Low confidence extraction for {ext.get('filename', 'unknown document')}. Manual review recommended."
                    )
                })
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict) and item.get('low_confidence'):
                        issues.append({
                            'severity': 'warning',
                            'category': 'Extraction Quality',
                            'description': item.get(
                                'confidence_warning',
                                f"Low confidence extraction for {ext.get('filename', 'unknown document')}. Manual review recommended."
                            )
                        })

    logger.info(f"Deterministic checks found {len(issues)} issues")
    return issues


def generate_issues(documents: List[Dict[str, Any]], cap_table: List[Dict[str, Any]], timeline: List[Dict[str, Any]], extractions: List[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """
    Generate an issue tracker by analyzing documents for gaps and inconsistencies.
    Combines deterministic (code-based) checks with AI analysis.

    Args:
        documents: List of classified documents
        cap_table: Generated cap table
        timeline: Generated timeline
        extractions: Raw extraction data (for deterministic checks)

    Returns:
        List of issues with severity and descriptions
    """
    # Start with deterministic issues (guaranteed to catch critical problems)
    deterministic_issues = []
    if extractions:
        deterministic_issues = check_deterministic_issues(documents, cap_table, timeline, extractions)

    try:
        # Prepare summaries for Claude
        doc_summary = [{'filename': d['filename'], 'category': d.get('category', 'Other')} for d in documents]

        documents_json = json.dumps(doc_summary, indent=2)
        cap_table_json = json.dumps(cap_table, indent=2)
        timeline_json = json.dumps(timeline, indent=2)

        prompt = prompts.ISSUE_TRACKER_PROMPT.format(
            documents_json=documents_json,
            cap_table_json=cap_table_json,
            timeline_json=timeline_json
        )

        # Add small delay to spread API load
        time.sleep(2)

        # Call Claude - if rate limited, return deterministic issues only (no retry)
        response = call_claude(prompt, max_tokens=4096)
        ai_issues = parse_json_response(response)

        # Combine deterministic (high confidence) with AI analysis (nuanced insights)
        all_issues = [normalize_issue(issue) for issue in (deterministic_issues + ai_issues)]
        logger.info(f"Total issues: {len(all_issues)} ({len(deterministic_issues)} deterministic, {len(ai_issues)} AI-detected)")

        return all_issues

    except RateLimitError as e:
        logger.warning(f"Issue generation rate limited, returning deterministic issues only: {e}")
        print(f"WARNING: AI issue analysis rate limited - showing {len(deterministic_issues)} deterministic issues")
        return [normalize_issue(issue) for issue in deterministic_issues]  # Return deterministic issues instead of empty/error
    except Exception as e:
        logger.error(f"AI issue generation failed: {e}", exc_info=True)
        print(f"ERROR: AI issue analysis failed - showing {len(deterministic_issues)} deterministic issues: {e}")
        # Return deterministic issues even if AI fails
        fallback = deterministic_issues if deterministic_issues else [
            {'severity': 'critical', 'category': 'System Error', 'description': f'Issue analysis failed: {str(e)}'}
        ]
        return [normalize_issue(issue) for issue in fallback]


def extract_company_name(extractions: List[Dict[str, Any]]) -> str:
    """
    Extract company name from charter documents.

    Args:
        extractions: List of documents with extracted data

    Returns:
        Company name string, or "Unknown Company" if not found
    """
    try:
        # Find charter documents
        charter_docs = [doc for doc in extractions if 'charter_data' in doc]

        if not charter_docs:
            return "Unknown Company"

        # Try to get company name from charter data
        for doc in charter_docs:
            charter_data = doc.get('charter_data', {})
            company_name = charter_data.get('company_name')
            if company_name:
                return company_name

        # If not found in structured data, ask Claude to extract it
        charter_texts = '\n\n---\n\n'.join([doc['text'][:5000] for doc in charter_docs if 'text' in doc])
        prompt = prompts.COMPANY_NAME_EXTRACTION_PROMPT.format(charter_texts=charter_texts)
        response = call_claude(prompt, max_tokens=256)

        return response.strip() or "Unknown Company"

    except Exception:
        return "Unknown Company"


# ============================================================================
# CAP TABLE TIE-OUT: Transaction Extraction & Approval Matching
# ============================================================================

TRANSACTABLE_DOC_TYPES = [
    'Stock Purchase Agreement',
    'SAFE',
    'Option Grant Agreement',
    'Share Repurchase Agreement',
    'Convertible Note'
]

APPROVAL_DOC_TYPES = [
    'Board/Shareholder Minutes',
    'Charter Document'  # For formation events
]


def extract_equity_transactions(extractions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Extract equity transactions from extracted data and structure for equity_events table.
    This is Pass 2A: Transaction extraction (atomic, like current MVP).

    Args:
        extractions: List of documents with extracted data from Pass 2

    Returns:
        List of transaction dictionaries ready for approval matching
    """
    transactions = []

    for doc in extractions:
        doc_id = doc.get('document_id')  # Set during document insertion
        filename = doc.get('filename', 'unknown')

        # Stock issuances
        if 'stock_issuances' in doc:
            for issuance in doc['stock_issuances']:
                if issuance.get('date') and issuance.get('shareholder') and issuance.get('shares'):
                    tx_details = {
                        'price_per_share': issuance.get('price_per_share'),
                        'verification': issuance.get('verification')
                    }
                    if doc.get('preview_focus_y') is not None:
                        tx_details['preview_focus_y'] = doc.get('preview_focus_y')

                    transactions.append({
                        'event_date': issuance['date'],
                        'event_type': 'issuance',
                        'shareholder_name': issuance['shareholder'],
                        'share_class': issuance.get('share_class', 'Common Stock'),
                        'share_delta': abs(issuance['shares']),  # Positive
                        'source_doc_id': doc_id,
                        'source_snippet': issuance.get('source_quote', f"{issuance['shareholder']} - {issuance['shares']} shares"),
                        'preview_image': doc.get('preview_image'),
                        'summary': doc.get('summary'),
                        'details': tx_details
                    })

        # SAFEs
        if 'safe_data' in doc:
            safe = doc['safe_data']
            if not safe.get('error') and safe.get('date') and safe.get('investor'):
                tx_details = {
                    'amount': safe.get('amount'),
                    'valuation_cap': safe.get('valuation_cap'),
                    'discount_rate': safe.get('discount_rate')
                }
                if doc.get('preview_focus_y') is not None:
                    tx_details['preview_focus_y'] = doc.get('preview_focus_y')

                transactions.append({
                    'event_date': safe['date'],
                    'event_type': 'safe',
                    'shareholder_name': safe['investor'],
                    'share_class': 'SAFE',
                    'share_delta': 0,  # SAFEs don't have shares yet
                    'source_doc_id': doc_id,
                    'source_snippet': safe.get('source_quote', f"SAFE investment by {safe['investor']}"),
                    'preview_image': doc.get('preview_image'),
                    'summary': doc.get('summary'),
                    'details': tx_details
                })

        # Convertible Notes
        if 'convertible_note_data' in doc:
            note = doc['convertible_note_data']
            if not note.get('error') and note.get('date') and note.get('investor'):
                tx_details = {
                    'principal': note.get('principal'),
                    'interest_rate': note.get('interest_rate'),
                    'maturity_date': note.get('maturity_date'),
                    'valuation_cap': note.get('valuation_cap'),
                    'discount_rate': note.get('discount_rate')
                }
                if doc.get('preview_focus_y') is not None:
                    tx_details['preview_focus_y'] = doc.get('preview_focus_y')

                transactions.append({
                    'event_date': note['date'],
                    'event_type': 'convertible_note',
                    'shareholder_name': note['investor'],
                    'share_class': 'Convertible Note',
                    'share_delta': 0,  # Notes don't have shares until conversion
                    'source_doc_id': doc_id,
                    'source_snippet': note.get('source_quote', f"Convertible note from {note['investor']}"),
                    'preview_image': doc.get('preview_image'),
                    'summary': doc.get('summary'),
                    'details': tx_details
                })

        # Option grants
        if 'option_data' in doc:
            option = doc['option_data']
            if not option.get('error') and option.get('grant_date') and option.get('recipient'):
                tx_details = {
                    'strike_price': option.get('strike_price'),
                    'vesting_schedule': option.get('vesting_schedule')
                }
                if doc.get('preview_focus_y') is not None:
                    tx_details['preview_focus_y'] = doc.get('preview_focus_y')

                transactions.append({
                    'event_date': option['grant_date'],
                    'event_type': 'option_grant',
                    'shareholder_name': option['recipient'],
                    'share_class': 'Option',
                    'share_delta': option.get('shares', 0),
                    'source_doc_id': doc_id,
                    'source_snippet': option.get('source_quote', f"Option grant to {option['recipient']}"),
                    'preview_image': doc.get('preview_image'),
                    'summary': doc.get('summary'),
                    'details': tx_details
                })

        # Repurchases (negative delta)
        if 'repurchase_data' in doc:
            repurchase = doc['repurchase_data']
            if not repurchase.get('error') and repurchase.get('date') and repurchase.get('shareholder'):
                shares = repurchase.get('shares')
                if shares and isinstance(shares, (int, float, Decimal)):
                    rep_details = {
                        'price_per_share': repurchase.get('price_per_share')
                    }
                    if doc.get('preview_focus_y') is not None:
                        rep_details['preview_focus_y'] = doc.get('preview_focus_y')

                    transactions.append({
                        'event_date': repurchase['date'],
                        'event_type': 'repurchase',
                        'shareholder_name': repurchase['shareholder'],
                        'share_class': repurchase.get('share_class', 'Common Stock'),
                        'share_delta': -float(abs(shares)),  # Negative for repurchases (convert to float)
                        'source_doc_id': doc_id,
                        'source_snippet': repurchase.get('source_quote', f"Repurchase from {repurchase['shareholder']}"),
                        'preview_image': doc.get('preview_image'),
                        'summary': doc.get('summary'),
                        'details': rep_details
                    })

        # Formation events from charter
        if 'charter_data' in doc:
            charter = doc['charter_data']
            if not charter.get('error') and charter.get('incorporation_date'):
                transactions.append({
                    'event_date': charter['incorporation_date'],
                    'event_type': 'formation',
                    'shareholder_name': None,
                    'share_class': None,
                    'share_delta': 0,
                    'source_doc_id': doc_id,
                    'source_snippet': charter.get('source_quote', f"Company incorporated: {charter.get('company_name')}"),
                    'details': {
                        'company_name': charter.get('company_name'),
                        'authorized_shares': charter.get('authorized_shares'),
                        'share_classes': charter.get('share_classes')
                    }
                })

    logger.info(f"Extracted {len(transactions)} equity transactions from documents")
    return transactions


def match_approvals_batch(transactions: List[Dict[str, Any]], classified_docs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Match transactions with their approving documents using a single batched Claude call.
    This is Pass 2B: Context-aware approval matching (batched, efficient).

    Args:
        transactions: List of transaction dicts from extract_equity_transactions()
        classified_docs: List of all classified documents (to find approval docs)

    Returns:
        Transactions with approval_doc_id, approval_snippet, compliance_status filled in
    """
    if not transactions:
        return transactions

    approval_required_types = {"issuance", "repurchase", "option_grant"}

    # Set conservative defaults before AI matching so unmatched rows never fail open.
    for tx in transactions:
        tx_type = (tx.get("event_type") or "").lower()
        tx["approval_doc_id"] = None
        tx["approval_snippet"] = None

        if tx_type in approval_required_types:
            tx["compliance_status"] = "WARNING"
            tx["compliance_note"] = "Approval evidence not matched automatically. Manual review required."
        elif tx_type == "formation":
            tx["compliance_status"] = "VERIFIED"
            tx["compliance_note"] = "Formation evidence sourced from charter document."
        else:
            tx["compliance_status"] = "WARNING"
            tx["compliance_note"] = "Approval linkage not required but evidence should be reviewed."

    # Filter to approval documents only
    approval_docs = [
        d for d in classified_docs
        if d.get('category') in APPROVAL_DOC_TYPES and d.get('document_id')
    ]

    if not approval_docs:
        # No approval documents found - mark approval-required tx as CRITICAL.
        logger.warning("No approval documents found - marking approval-required transactions as CRITICAL")
        for tx in transactions:
            tx_type = (tx.get("event_type") or "").lower()
            if tx_type in approval_required_types:
                tx['compliance_status'] = 'CRITICAL'
                tx['compliance_note'] = 'No board approval documents found in document set'
            elif tx_type == "formation":
                tx['compliance_status'] = 'VERIFIED'
                tx['compliance_note'] = 'Formation evidence sourced from charter document'
            else:
                tx['compliance_status'] = 'WARNING'
                tx['compliance_note'] = 'No approval documents found to validate this financing event'
        return transactions

    # Build manifest for Claude (lightweight, just metadata + excerpt)
    manifest = []
    for doc in approval_docs:
        manifest.append({
            'doc_id': str(doc['document_id']),
            'filename': doc.get('filename', 'unknown'),
            'category': doc.get('category'),
            'excerpt': doc.get('text', '')[:2000] if doc.get('text') else ''  # First 2K chars
        })

    # Build transaction summary for Claude
    tx_summary = []
    for i, tx in enumerate(transactions):
        tx_summary.append({
            'tx_index': i,
            'event_date': tx['event_date'],
            'event_type': tx['event_type'],
            'shareholder': tx.get('shareholder_name'),
            'shares': tx.get('share_delta'),
            'snippet': (tx.get('source_snippet') or '')[:500]  # Truncate for API efficiency
        })

    # Single Claude call for batch approval matching
    try:
        prompt = prompts.BATCH_APPROVAL_MATCHING_PROMPT.format(
            transactions_json=json.dumps(tx_summary, indent=2),
            approval_docs_json=json.dumps(manifest, indent=2)
        )

        logger.info(f"Matching {len(transactions)} transactions with {len(approval_docs)} approval documents via batch Claude call...")
        response = call_claude(prompt, max_tokens=8000)
        matches = parse_json_response(response)

        # Validate response structure
        if not isinstance(matches, list):
            raise ValueError(f"Expected JSON array, got {type(matches).__name__}")
        if matches and not all('tx_index' in m for m in matches):
            logger.warning("Some matches missing tx_index field - filtering out invalid entries")
            matches = [m for m in matches if 'tx_index' in m]

        # Build set of valid approval doc IDs for validation
        valid_approval_ids = {str(doc['document_id']) for doc in approval_docs if doc.get('document_id')}

        # Merge results back into transactions
        for match in matches:
            tx_idx = match.get('tx_index')
            if tx_idx is not None and 0 <= tx_idx < len(transactions):
                returned_id = match.get('approval_doc_id')

                if returned_id and str(returned_id) not in valid_approval_ids:
                    logger.warning(
                        f"Claude returned invalid approval_doc_id '{returned_id}' for tx_index {tx_idx} "
                        f"(not in approval manifest of {len(valid_approval_ids)} docs). Setting to null."
                    )
                    transactions[tx_idx]['approval_doc_id'] = None
                    transactions[tx_idx]['approval_snippet'] = None
                    transactions[tx_idx]['compliance_status'] = 'WARNING'
                    transactions[tx_idx]['compliance_note'] = (
                        (match.get('compliance_note') or '') +
                        ' [Auto-corrected: AI returned non-approval document reference]'
                    ).strip()
                else:
                    transactions[tx_idx]['approval_doc_id'] = returned_id
                    transactions[tx_idx]['approval_snippet'] = match.get('approval_quote')
                    fallback_status = "WARNING"
                    if (transactions[tx_idx].get('event_type') or '').lower() == 'formation':
                        fallback_status = "VERIFIED"
                    transactions[tx_idx]['compliance_status'] = _normalize_compliance_status(
                        match.get('compliance_status'),
                        fallback=fallback_status
                    )
                    transactions[tx_idx]['compliance_note'] = match.get('compliance_note')

        logger.info(f"Batch approval matching complete: {len(matches)} matches processed")
        return transactions

    except Exception as e:
        logger.error(f"Batch approval matching failed: {e}", exc_info=True)
        # Fallback defaults already applied above.
        for tx in transactions:
            tx['compliance_note'] = f'Approval matching failed: {str(e)}'
        return transactions


# ============================================================================
# MAIN ORCHESTRATOR
# ============================================================================

def process_audit(audit_id: str, documents: List[Dict[str, Any]]):
    """
    Main orchestrator for the 3-pass AI audit pipeline.

    Args:
        audit_id: UUID of the audit
        documents: List of parsed documents with 'filename', 'type', 'text' fields
    """
    try:
        total_docs = len(documents)
        transactions: List[Dict[str, Any]] = []
        logger.info(f"Starting 3-pass processing for {total_docs} documents")

        # ========== PASS 1: CLASSIFICATION ==========
        try:
            db.update_progress(audit_id, "Pass 1: Classifying documents...", pipeline_state='classifying')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")
            print(f"ERROR: Failed to update progress: {e}")

        classified_docs = []
        for i, doc in enumerate(documents, start=1):
            try:
                db.update_progress(
                    audit_id,
                    f"Pass 1: Classifying documents... {i}/{total_docs}",
                    pipeline_state='classifying'
                )
            except Exception as e:
                logger.error(f"Failed to update progress: {e}")

            logger.info(f"Classifying document {i}/{total_docs}: {doc.get('filename', 'unknown')}")
            classified_doc = classify_document(doc)
            classified_doc['parse_status'] = doc.get('parse_status', 'error' if doc.get('error') else 'success')
            classified_doc['parse_error'] = doc.get('parse_error') or doc.get('error')
            classified_docs.append(classified_doc)

        logger.info(f"Pass 1 complete: Classified {total_docs} documents")

        # ========== PASS 2: EXTRACTION ==========
        try:
            db.update_progress(audit_id, "Pass 2: Extracting structured data...", pipeline_state='extracting')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")
            print(f"ERROR: Failed to update progress: {e}")

        extractions = []
        for i, doc in enumerate(classified_docs, start=1):
            try:
                db.update_progress(
                    audit_id,
                    f"Pass 2: Extracting data... {i}/{total_docs}",
                    pipeline_state='extracting'
                )
            except Exception as e:
                logger.error(f"Failed to update progress: {e}")

            if doc.get('error') or doc.get('parse_status') == 'error':
                extractions.append(doc)
                continue

            extracted_data = extract_by_type(doc)
            extractions.append({**doc, **extracted_data})

        logger.info(f"Pass 2 complete: Extracted data from {total_docs} documents")

        # Build versioned extracted_data payloads for each persisted document.
        enriched_docs = _build_enriched_documents(extractions)

        # Persist normalized documents first so transaction records can reference document IDs.
        try:
            db.update_progress(audit_id, "Pass 2: Saving normalized documents...", pipeline_state='extracting')
            doc_ids = db.insert_documents_and_events(audit_id, enriched_docs, [])
            for doc in extractions:
                doc_key = str(doc.get('id') or doc.get('filename') or 'unknown')
                if doc_key in doc_ids:
                    doc['document_id'] = doc_ids[doc_key]
            for doc in enriched_docs:
                doc_key = str(doc.get('id') or doc.get('filename') or 'unknown')
                if doc_key in doc_ids:
                    doc['document_id'] = doc_ids[doc_key]
            logger.info(f"Inserted {len(enriched_docs)} documents")
        except Exception as e:
            logger.error(f"Failed to insert normalized documents: {e}", exc_info=True)
            raise

        # ========== PASS 2 TIE-OUT: Transaction Extraction & Approval Matching ==========
        try:
            db.update_progress(audit_id, "Pass 2: Matching transactions with approvals...", pipeline_state='reconciling')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")

        transactions = extract_equity_transactions(extractions)

        if transactions:
            transactions = match_approvals_batch(transactions, extractions)

            # Enforce stable schema for equity_events.details payloads.
            for tx in transactions:
                details = tx.get('details') if isinstance(tx.get('details'), dict) else {}
                tx['details'] = {'schema_version': 'v1', **details}

            db.insert_equity_events(audit_id, transactions)
            logger.info(f"Inserted {len(transactions)} equity events")

        # ========== PASS 3: SYNTHESIS ==========
        try:
            db.update_progress(audit_id, "Pass 3: Synthesizing timeline...", pipeline_state='reconciling')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")
            print(f"ERROR: Failed to update progress: {e}")

        timeline = synthesize_timeline(extractions)

        try:
            db.update_progress(audit_id, "Pass 3: Synthesizing cap table...", pipeline_state='reconciling')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")

        cap_table, cap_table_issues = synthesize_cap_table(extractions)

        try:
            db.update_progress(audit_id, "Pass 3: Synthesizing issues...", pipeline_state='reconciling')
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")

        issues = generate_issues(enriched_docs, cap_table, timeline, extractions)
        issues.extend([normalize_issue(issue) for issue in cap_table_issues])
        issues = [normalize_issue(issue) for issue in issues]

        company_name = extract_company_name(extractions)
        logger.info("Pass 3 complete: Generated timeline, cap table, and issues")

        # ========== SAVE RESULTS ==========
        failed_docs = [
            d for d in enriched_docs
            if d.get('parse_status') == 'error' or d.get('error')
        ]

        cleaned_docs = [clean_document_dict(doc) for doc in enriched_docs]
        cleaned_failed_docs = [clean_document_dict(doc) for doc in failed_docs]
        quality_report = build_quality_report(cleaned_docs, transactions, issues)
        review_required = bool(quality_report.get('review_required'))

        try:
            db.update_progress(
                audit_id,
                "Manual review required before finalization." if review_required else "Pass 3: Finalizing report...",
                pipeline_state='needs_review' if review_required else 'complete'
            )
        except Exception as e:
            logger.error(f"Failed to update progress: {e}")

        db.update_audit_results(
            audit_id,
            {
                'company_name': company_name,
                'documents': cleaned_docs,
                'timeline': timeline,
                'cap_table': cap_table,
                'issues': issues,
                'failed_documents': cleaned_failed_docs
            },
            review_required=review_required,
            quality_report=quality_report
        )

        logger.info(
            f"Audit {audit_id} completed with status "
            f"{'needs_review' if review_required else 'complete'}"
        )

    except Exception as e:
        logger.error(f"Processing error in audit {audit_id}: {e}", exc_info=True)
        print(f"ERROR: Processing failed for audit {audit_id}: {e}")
        try:
            db.mark_error(audit_id, str(e))
        except Exception as db_error:
            logger.error(f"Failed to mark error in database: {db_error}")
            print(f"CRITICAL: Failed to mark error in database: {db_error}")


# ============================================================================
# CARTA CAP TABLE TIE-OUT
# ============================================================================

def parse_carta_captable(xlsx_path: str) -> Dict[str, Any]:
    """
    Parse a Carta-exported cap table (.xlsx) into structured data.
    Handles multiple Carta export formats (Summary, Detailed, Certificate Ledger).

    Returns:
        Dictionary with shareholders list and total_shares
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    shareholders = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Find header row (scan first 5 rows for recognizable headers)
        header_row = None
        headers = []
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_vals = [str(cell.value).lower().strip() if cell.value else '' for cell in ws[row_idx]]
            if any(kw in h for h in row_vals for kw in ['name', 'stakeholder']):
                header_row = row_idx
                headers = row_vals
                break

        if header_row is None:
            continue

        # Detect column indices
        name_col = None
        shares_col = None
        ownership_col = None

        for i, h in enumerate(headers):
            if name_col is None and any(kw in h for kw in ['stakeholder name', 'name']):
                name_col = i
            if shares_col is None and any(kw in h for kw in ['outstanding shares', 'quantity outstanding', 'shares outstanding']):
                shares_col = i
            if ownership_col is None and any(kw in h for kw in ['outstanding ownership', 'fully diluted ownership']):
                ownership_col = i

        # Fallback: look for common stock column like "common (cs)"
        if shares_col is None:
            for i, h in enumerate(headers):
                if 'common' in h and ('cs' in h or 'stock' in h):
                    shares_col = i
                    break

        if name_col is None or shares_col is None:
            continue

        # Extract data rows
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            name = row[name_col] if name_col < len(row) else None
            shares_val = row[shares_col] if shares_col < len(row) else None

            if not name or not shares_val:
                continue

            name_str = str(name).strip()
            # Skip total/summary rows
            if name_str.lower() in ['total', 'totals', 'grand total', '']:
                continue
            if 'total' in name_str.lower() and ('issued' in name_str.lower() or 'outstanding' in name_str.lower()):
                continue
            # Skip non-stakeholder rows
            if 'options' in name_str.lower() or 'shares available' in name_str.lower():
                continue
            if 'fully diluted' in name_str.lower() or 'percentage' in name_str.lower():
                continue

            # Parse shares
            try:
                if isinstance(shares_val, str):
                    shares = float(shares_val.replace(',', '').replace('$', ''))
                else:
                    shares = float(shares_val)
            except (ValueError, TypeError):
                continue

            if shares <= 0:
                continue

            # Parse ownership if available
            ownership = 0.0
            if ownership_col and ownership_col < len(row) and row[ownership_col]:
                try:
                    own_val = str(row[ownership_col]).replace('%', '')
                    ownership = float(own_val)
                except (ValueError, TypeError):
                    pass

            shareholders.append({
                'name': name_str,
                'shares': shares,
                'share_class': 'Common Stock',
                'ownership_pct': ownership
            })

        if shareholders:
            break  # Found valid data, stop searching sheets

    # Aggregate by name (Carta may have multiple certificates per person)
    aggregated = {}
    for sh in shareholders:
        key = sh['name']
        if key not in aggregated:
            aggregated[key] = {'name': key, 'shares': 0.0, 'share_class': sh['share_class'], 'ownership_pct': 0.0}
        aggregated[key]['shares'] += sh['shares']

    shareholders = list(aggregated.values())
    total_shares = sum(sh['shares'] for sh in shareholders)

    # Recalculate ownership percentages
    if total_shares > 0:
        for sh in shareholders:
            sh['ownership_pct'] = round((sh['shares'] / total_shares) * 100, 2)

    return {
        'shareholders': shareholders,
        'total_shares': total_shares
    }


def _normalize_name(name: str) -> str:
    """Normalize a shareholder name for comparison."""
    if not name:
        return ''
    return name.lower().strip().replace(',', '').replace('.', '')


def _find_best_match(name: str, candidates) -> str:
    """
    Find best matching name from candidates.
    Conservative matching to avoid false positives in legal tie-outs.
    """
    import difflib

    if not name:
        return None
    candidates = list(candidates)
    if not candidates:
        return None

    # Exact match
    if name in candidates:
        return name

    # Fuzzy match only when confidence is high and unambiguous.
    scored = [
        (candidate, difflib.SequenceMatcher(None, name, candidate).ratio())
        for candidate in candidates
    ]
    scored.sort(key=lambda item: item[1], reverse=True)
    best_name, best_score = scored[0]
    second_score = scored[1][1] if len(scored) > 1 else 0.0

    if best_score >= 0.92 and (best_score - second_score) >= 0.05:
        return best_name

    return None


def compare_cap_tables(
    carta_shareholders: List[Dict[str, Any]],
    generated_cap_table: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Compare Carta cap table against pipeline-generated cap table.
    Returns list of discrepancy issues.
    """
    issues = []

    # Build normalized-name lookups with aggregation to avoid overwrite on duplicates.
    gen_by_name = {}
    gen_duplicates = set()
    for entry in generated_cap_table:
        name = _normalize_name(entry.get('shareholder', ''))
        if not name:
            continue
        shares = float(entry.get('shares', 0) or 0)
        if name in gen_by_name:
            gen_duplicates.add(name)
            gen_by_name[name]['shares'] = float(gen_by_name[name].get('shares', 0) or 0) + shares
        else:
            gen_by_name[name] = {
                **entry,
                'shares': shares
            }

    carta_by_name = {}
    carta_duplicates = set()
    for entry in carta_shareholders:
        name = _normalize_name(entry.get('name', ''))
        if not name:
            continue
        shares = float(entry.get('shares', 0) or 0)
        if name in carta_by_name:
            carta_duplicates.add(name)
            carta_by_name[name]['shares'] = float(carta_by_name[name].get('shares', 0) or 0) + shares
        else:
            carta_by_name[name] = {
                **entry,
                'shares': shares
            }

    for dup_name in sorted(gen_duplicates):
        issues.append({
            'severity': 'warning',
            'category': 'Cap Table Tie-Out',
            'description': f'Generated cap table has duplicate shareholder entries for "{gen_by_name[dup_name].get("shareholder", dup_name)}". Shares were aggregated before matching.'
        })

    for dup_name in sorted(carta_duplicates):
        issues.append({
            'severity': 'warning',
            'category': 'Cap Table Tie-Out',
            'description': f'Carta cap table has duplicate shareholder entries for "{carta_by_name[dup_name].get("name", dup_name)}". Shares were aggregated before matching.'
        })

    matched_gen_names = set()

    # Check each Carta shareholder against generated cap table
    for carta_name, carta_entry in carta_by_name.items():
        match = _find_best_match(carta_name, gen_by_name.keys())

        if match is None:
            issues.append({
                'severity': 'warning',
                'category': 'Cap Table Tie-Out',
                'description': (
                    f'Shareholder "{carta_entry["name"]}" appears in Carta cap table '
                    f'({carta_entry["shares"]:,.0f} shares) but was not found in source documents.'
                )
            })
            continue

        matched_gen_names.add(match)
        gen_entry = gen_by_name[match]
        carta_shares = carta_entry['shares']
        gen_shares = float(gen_entry.get('shares', 0))

        # Compare share counts (tolerance of 0.5 for rounding)
        if abs(carta_shares - gen_shares) > 0.5:
            issues.append({
                'severity': 'critical',
                'category': 'Cap Table Tie-Out',
                'description': (
                    f'Share count mismatch for "{carta_entry["name"]}": '
                    f'Carta shows {carta_shares:,.0f} shares, '
                    f'source documents show {gen_shares:,.0f} shares. '
                    f'Discrepancy: {abs(carta_shares - gen_shares):,.0f} shares.'
                )
            })

    # Check for shareholders in source documents but not in Carta
    for gen_name, gen_entry in gen_by_name.items():
        if gen_name not in matched_gen_names:
            issues.append({
                'severity': 'warning',
                'category': 'Cap Table Tie-Out',
                'description': (
                    f'Shareholder "{gen_entry["shareholder"]}" found in source documents '
                    f'({float(gen_entry.get("shares", 0)):,.0f} shares) but not in Carta cap table.'
                )
            })

    # Compare total shares
    carta_total = sum(sh['shares'] for sh in carta_shareholders)
    gen_total = sum(float(e.get('shares', 0)) for e in generated_cap_table)
    if abs(carta_total - gen_total) > 0.5:
        issues.append({
            'severity': 'critical',
            'category': 'Cap Table Tie-Out',
            'description': (
                f'Total share count mismatch: Carta shows {carta_total:,.0f} total shares, '
                f'source documents show {gen_total:,.0f} total shares. '
                f'Discrepancy: {abs(carta_total - gen_total):,.0f} shares.'
            )
        })

    return issues


def tieout_carta_captable(audit_id: str, captable_path: str):
    """
    Compare uploaded Carta cap table against the generated cap table.
    Appends discrepancies as issues to the audit.
    """
    # Parse Carta cap table
    carta_data = parse_carta_captable(captable_path)

    if not carta_data['shareholders']:
        logger.warning(f"Could not parse Carta cap table for audit {audit_id}")
        db.append_issues(audit_id, [normalize_issue({
            'severity': 'warning',
            'category': 'Cap Table Tie-Out',
            'description': 'Uploaded cap table could not be parsed. Ensure it is a standard Carta export (.xlsx).'
        })])
        return

    # Fetch the generated cap table
    audit = db.get_audit(audit_id)
    generated_cap_table = audit.get('cap_table', []) if audit else []

    if not generated_cap_table:
        logger.warning(f"No generated cap table to compare against for audit {audit_id}")
        return

    # Compare and generate issues
    issues = compare_cap_tables(carta_data['shareholders'], generated_cap_table)

    if issues:
        db.append_issues(audit_id, [normalize_issue(issue) for issue in issues])
        logger.info(f"Cap table tie-out found {len(issues)} discrepancies for audit {audit_id}")
    else:
        db.append_issues(audit_id, [normalize_issue({
            'severity': 'info',
            'category': 'Cap Table Tie-Out',
            'description': f'Carta cap table matches source documents. All {len(carta_data["shareholders"])} shareholders verified.'
        })])
